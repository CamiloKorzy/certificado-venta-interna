from fastapi import FastAPI, HTTPException, Request, Depends, UploadFile, File, Form
import io
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import psycopg2
from datetime import datetime
import os
import openpyxl
import difflib
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import sys
import traceback

# INYECCIÓN CRÍTICA PARA VERCEL: los módulos están en /var/task/backend
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

from auth import hash_password, verify_password, create_token, decode_token
from telegram_service import telegram_nuevo_certificado, telegram_test


def safe_float(val):
    if not val:
        return 0.0
    if isinstance(val, str) and val.strip().upper() == 'NULL':
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def normalize_periodo(periodo: str) -> str:
    if not periodo:
        return periodo
    periodo = str(periodo).replace("/", "-").strip()
    parts = periodo.split('-')
    if len(parts) == 2:
        if len(parts[0]) == 4:
            return f"{parts[0]}-{parts[1].zfill(2)}"
        elif len(parts[1]) == 4:
            return f"{parts[1]}-{parts[0].zfill(2)}"
    return periodo

def check_informe_cerrado(empresa: str, periodo: str, modulo: str):
    try:
        periodo = normalize_periodo(periodo)
        if not empresa or not periodo: return None
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("SELECT estado, snapshot_data FROM cert_informes_proyecto WHERE unidad_negocio = %s AND periodo = %s", (empresa, periodo))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row[0] == 'CERRADO' and row[1]:
            data = row[1].get(modulo)
            if data and isinstance(data, list):
                for item in data:
                    if isinstance(item, dict) and "categoria" in item:
                        db_cat = item.get("categoria") or "Ajuste Manual"
                        if db_cat == "Ajuste Excel":
                            item["categoria"] = "Ingresos adicionales" if item.get("tipo_movimiento") == "INGRESO" else "Egresos adicionales"
                        elif db_cat == "Gastos de Compra":
                            item["categoria"] = "Costos"
            return data
    except:
        pass
    return None

app = FastAPI(title="API Certificado de Ventas Internos - CEE ENRIQUEZ")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Middleware que atrapa TODAS las excepciones y devuelve JSON informativo
from fastapi.responses import JSONResponse

@app.middleware("http")
async def catch_exceptions_middleware(request: Request, call_next):
    try:
        response = await call_next(request)
        return response
    except Exception as e:
        tb = traceback.format_exc()
        print(f"UNHANDLED EXCEPTION: {tb}")
        return JSONResponse(status_code=500, content={"error": str(e), "trace": tb[:500]})

# ─── Credenciales ───
DB_HOST = os.environ.get("DB_HOST", "ceesa.dw.finneg.com")
DB_PORT = os.environ.get("DB_PORT", "5432")
DB_NAME = os.environ.get("DB_NAME", "finnegansbi")
DB_USER = os.environ.get("DB_USER", "ceesauser")
DB_PASS = os.environ.get("DB_PASS", "Lula$$2014")
SUPABASE_DB_URL = os.environ.get("SUPABASE_DB_URL", "postgresql://postgres:FsIPyXBJT8aFZk8D@db.rsofgomdfrrvawvqybxp.supabase.co:5432/postgres")
APP_URL = os.environ.get("APP_URL", "https://certificado-venta-interna.vercel.app")

def get_aurora():
    return psycopg2.connect(host=DB_HOST, port=DB_PORT, database=DB_NAME, user=DB_USER, password=DB_PASS, sslmode="require")

import socket

# PARCHE PARA VERCEL: Forzar IPv4 (AF_INET)
# Vercel intenta usar IPv6 por defecto para Supabase, pero falla con "Cannot assign requested address".
old_getaddrinfo = socket.getaddrinfo
def new_getaddrinfo(*args, **kwargs):
    responses = old_getaddrinfo(*args, **kwargs)
    return [r for r in responses if r[0] == socket.AF_INET]
socket.getaddrinfo = new_getaddrinfo

def get_supabase():
    if not SUPABASE_DB_URL:
        raise HTTPException(status_code=500, detail="SUPABASE_DB_URL no configurada")
    try:
        return psycopg2.connect(SUPABASE_DB_URL, connect_timeout=10)
    except Exception as e:
        if "Cannot assign requested address" in str(e) or "IPv6" in str(e):
            raise HTTPException(
                status_code=500, 
                detail="Error de IPv6 en Vercel. Por favor, entrá a Supabase -> Project Settings -> Database y copiá la URL del 'Connection Pooler' (puerto 6543) y reemplazá SUPABASE_DB_URL en Vercel."
            )
        raise e

# ─── Auto-setup: crea tablas en Supabase si no existen ───
_setup_done = False
def auto_setup_db():
    global _setup_done
    if _setup_done:
        return
    _setup_done = True
    if not SUPABASE_DB_URL:
        print("[SETUP] SUPABASE_DB_URL no configurada, saltando auto-setup")
        return
    try:
        conn = psycopg2.connect(SUPABASE_DB_URL, connect_timeout=10)
        conn.autocommit = True
        cur = conn.cursor()
        
        # Migración: Si las tablas de configuración avanzada ya existen pero no tienen la columna 'sucursal', las eliminamos para recrearlas.
        try:
            cur.execute("""
                SELECT EXISTS (
                    SELECT 1 
                    FROM information_schema.columns 
                    WHERE table_name = 'cert_config_ingresos_comprobantes' 
                      AND column_name = 'sucursal'
                )
            """)
            has_sucursal = cur.fetchone()[0]
            
            # También verificamos si la tabla existe. Si no existe, has_sucursal será False.
            cur.execute("SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = 'cert_config_ingresos_comprobantes')")
            table_exists = cur.fetchone()[0]
            
            if table_exists and not has_sucursal:
                print("[MIGRACIÓN] Estructura global detectada. Eliminando tablas antiguas de configuración avanzada para recrearlas por sucursal...")
                cur.execute("DROP TABLE IF EXISTS cert_config_ingresos_comprobantes")
                cur.execute("DROP TABLE IF EXISTS cert_config_gastos_asientos")
                cur.execute("DROP TABLE IF EXISTS cert_config_gastos_compras")
        except Exception as e:
            print("[MIGRACIÓN] Error durante la comprobación de migración:", e)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS cert_usuarios (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                password TEXT NOT NULL,
                rol TEXT NOT NULL,
                telegram_chat_id TEXT,
                activo INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS cert_usuarios_unidades (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER REFERENCES cert_usuarios(id) ON DELETE CASCADE,
                unidad_negocio TEXT NOT NULL,
                notifica_email BOOLEAN DEFAULT false,
                notifica_telegram BOOLEAN DEFAULT false,
                UNIQUE(usuario_id, unidad_negocio)
            );
            CREATE TABLE IF NOT EXISTS cert_notificaciones_log (
                id SERIAL PRIMARY KEY,
                tipo TEXT,
                destinatario TEXT,
                comprobante TEXT,
                mensaje TEXT,
                estado TEXT,
                error TEXT,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_cert_usuarios_email ON cert_usuarios(email);
            CREATE INDEX IF NOT EXISTS idx_cert_usu_unid_usuario ON cert_usuarios_unidades(usuario_id);
            CREATE INDEX IF NOT EXISTS idx_cert_notif_fecha ON cert_notificaciones_log(fecha);
            
            CREATE TABLE IF NOT EXISTS cert_audit_log (
                id SERIAL PRIMARY KEY,
                usuario_email TEXT,
                accion TEXT,
                detalles TEXT,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_cert_audit_fecha ON cert_audit_log(fecha);

            CREATE TABLE IF NOT EXISTS cert_ajustes_excel (
                id SERIAL PRIMARY KEY,
                unidad_negocio VARCHAR(255) NOT NULL,
                periodo VARCHAR(7) NOT NULL,
                concepto VARCHAR(255) NOT NULL,
                tipo_movimiento VARCHAR(50) NOT NULL,
                categoria VARCHAR(255),
                importe NUMERIC(15, 2) NOT NULL,
                observaciones TEXT,
                fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                usuario_carga VARCHAR(255)
            );
            
            CREATE TABLE IF NOT EXISTS cert_config_gastos_cuentas (
                id SERIAL PRIMARY KEY,
                categoria TEXT NOT NULL,
                cuenta_codigo TEXT NOT NULL,
                cuenta_nombre TEXT NOT NULL,
                UNIQUE(categoria, cuenta_codigo)
            );
            CREATE INDEX IF NOT EXISTS idx_cert_config_gastos_cat ON cert_config_gastos_cuentas(categoria);
            
            CREATE TABLE IF NOT EXISTS cert_config_ingresos_comprobantes (
                id SERIAL PRIMARY KEY,
                sucursal TEXT NOT NULL,
                subtipo_id TEXT NOT NULL,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                UNIQUE(sucursal, subtipo_id)
            );
            
            CREATE TABLE IF NOT EXISTS cert_config_gastos_asientos (
                id SERIAL PRIMARY KEY,
                sucursal TEXT NOT NULL,
                tipo_asiento_id TEXT NOT NULL,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                UNIQUE(sucursal, tipo_asiento_id)
            );
            
            CREATE TABLE IF NOT EXISTS cert_config_gastos_compras (
                id SERIAL PRIMARY KEY,
                sucursal TEXT NOT NULL,
                subtipo_id TEXT NOT NULL,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                UNIQUE(sucursal, subtipo_id)
            );
            
            CREATE TABLE IF NOT EXISTS cert_config_centros_costo (
                id SERIAL PRIMARY KEY,
                sucursal TEXT NOT NULL,
                centro_id TEXT NOT NULL,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                UNIQUE(sucursal, centro_id)
            );

            
            CREATE TABLE IF NOT EXISTS cert_excel_uploads (
                id SERIAL PRIMARY KEY,
                nombre_archivo TEXT NOT NULL,
                usuario TEXT NOT NULL,
                periodo TEXT,
                fecha_subida TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            
            CREATE TABLE IF NOT EXISTS cert_excel_items (
                id SERIAL PRIMARY KEY,
                upload_id INTEGER REFERENCES cert_excel_uploads(id) ON DELETE CASCADE,
                unidad_negocio TEXT NOT NULL,
                fecha DATE NOT NULL,
                concepto TEXT NOT NULL,
                tipo TEXT NOT NULL,
                importe NUMERIC(15,2) NOT NULL
            );
            
            CREATE TABLE IF NOT EXISTS cert_cierres_mensuales (
                id SERIAL PRIMARY KEY,
                unidad_negocio TEXT NOT NULL,
                periodo TEXT NOT NULL,
                usuario_cierre TEXT NOT NULL,
                fecha_cierre TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(unidad_negocio, periodo)
            );
            
            CREATE TABLE IF NOT EXISTS cert_cierres_detalle (
                id SERIAL PRIMARY KEY,
                cierre_id INTEGER REFERENCES cert_cierres_mensuales(id) ON DELETE CASCADE,
                origen TEXT NOT NULL,
                tipo_movimiento TEXT NOT NULL,
                categoria TEXT,
                fecha DATE,
                concepto TEXT,
                comprobante TEXT,
                proveedor TEXT,
                importe NUMERIC(15,2) NOT NULL
            );
            
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name='cert_cierres_detalle' AND column_name='proveedor') THEN
                    ALTER TABLE cert_cierres_detalle ADD COLUMN proveedor TEXT;
                END IF;
                
                CREATE TABLE IF NOT EXISTS cert_informes_proyecto (
                    id SERIAL PRIMARY KEY,
                    unidad_negocio TEXT NOT NULL,
                    periodo TEXT NOT NULL,
                    estado TEXT NOT NULL DEFAULT 'ABIERTO',
                    snapshot_data JSONB,
                    usuario_apertura TEXT,
                    fecha_apertura TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    usuario_cierre TEXT,
                    fecha_cierre TIMESTAMP,
                    UNIQUE(unidad_negocio, periodo)
                );
            END $$;

            CREATE TABLE IF NOT EXISTS cert_respaldos (
                id SERIAL PRIMARY KEY,
                tipo_documento VARCHAR(100) NOT NULL,
                unidad_negocio VARCHAR(255) NOT NULL,
                periodo VARCHAR(7) NOT NULL,
                nombre_archivo VARCHAR(255) NOT NULL,
                tipo_mime VARCHAR(255),
                contenido TEXT,
                usuario_carga VARCHAR(255),
                fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS cert_config_equipos_asignados (
                id SERIAL PRIMARY KEY,
                sucursal TEXT NOT NULL,
                equipo_codigo TEXT NOT NULL,
                equipo_nombre TEXT NOT NULL,
                fecha_desde DATE NOT NULL,
                fecha_hasta DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                usuario_carga TEXT,
                fecha_carga TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX IF NOT EXISTS idx_cert_config_equipos_suc ON cert_config_equipos_asignados(sucursal);

        """)
        # Crear admin por defecto si no existe
        cur.execute("SELECT COUNT(*) FROM cert_usuarios WHERE email = 'admin@ceeenriquez.com'")
        if cur.fetchone()[0] == 0:
            admin_hash = hash_password("admin2026")
            cur.execute("""
                INSERT INTO cert_usuarios (email, nombre, password, rol, activo)
                VALUES ('admin@ceeenriquez.com', 'Administrador', %s, 'admin', 1)
            """, (admin_hash,))
            print("[SETUP] OK - Usuario admin creado: admin@ceeenriquez.com / admin2026")
            
        # Migración: Agregar columna estado_doc si no existe
        try:
            cur.execute("ALTER TABLE cert_notificaciones_log ADD COLUMN IF NOT EXISTS estado_doc TEXT;")
        except Exception as e:
            pass
            
        print("[SETUP] OK - Tablas cert_* verificadas/creadas")
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[SETUP] ERROR - Error en auto-setup: {e}")
        _setup_done = False  # Reintentar en próxima llamada

def log_action(email: str, accion: str, detalles: str = ""):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO cert_audit_log (usuario_email, accion, detalles) VALUES (%s, %s, %s)",
            (email, accion, detalles)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[AUDIT LOG ERROR] {e}")

# ─── Auth Dependency ───
async def get_current_user(request: Request):
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token requerido")
    token = auth_header.split(" ")[1]
    payload = decode_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Token inválido o expirado")
    return payload

async def require_admin(user=Depends(get_current_user)):
    if user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Acceso denegado: se requiere rol admin")
    return user

# ─── Pydantic Models ───
class LoginRequest(BaseModel):
    email: str
    password: str

class UsuarioCreate(BaseModel):
    nombre: str
    email: str
    password: str
    rol: str = "consulta"
    telegram_chat_id: Optional[str] = None
    activo: int = 1

class UsuarioUpdate(BaseModel):
    nombre: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    rol: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    activo: Optional[int] = None

class UnidadAsignacion(BaseModel):
    unidad_negocio: str

# ═══════════════════════════════════════════════════════
# ENDPOINTS DE AUTENTICACIÓN
# ═══════════════════════════════════════════════════════

@app.get("/api/health")
def health_check():
    import traceback
    try:
        info = {"status": "ok", "supabase_url_set": bool(SUPABASE_DB_URL), "setup_done": _setup_done}
        info["supabase_url_preview"] = SUPABASE_DB_URL[:30] + "..." if SUPABASE_DB_URL else "VACIA"
        # Test Supabase connectivity
        try:
            conn = psycopg2.connect(SUPABASE_DB_URL, connect_timeout=5)
            cur = conn.cursor()
            cur.execute("SELECT 1")
            info["supabase_connect"] = "OK"
            try:
                cur.execute("SELECT COUNT(*) FROM cert_usuarios")
                count = cur.fetchone()[0]
                info["users_count"] = count
            except Exception as e2:
                info["table_check"] = f"tabla no existe: {str(e2)[:100]}"
            cur.close()
            conn.close()
        except Exception as e:
            info["supabase_connect"] = f"error: {str(e)[:200]}"
        return info
    except Exception as e:
        return {"error": str(e), "trace": traceback.format_exc()[:500]}

@app.get("/api/setup")
def run_setup():
    auto_setup_db()
    return {"status": "ok", "message": "Base de datos (Supabase) inicializada correctamente."}

# ═══════════════════════════════════════════════════════
# ENDPOINTS CONFIGURACIÓN GASTOS
# ═══════════════════════════════════════════════════════

@app.get("/api/config/cuentas")
def get_config_cuentas():
    """Obtiene el listado de cuentas de resultado de FinnegansBI para configurar gastos."""
    try:
        conn = get_aurora()
        cur = conn.cursor()
        cur.execute("""
            SELECT CuentaID, Codigo, Nombre
            FROM BSCuenta
            WHERE ImpactaResultados = 1
            ORDER BY Nombre;
        """)
        cuentas = []
        for row in cur.fetchall():
            cuentas.append({
                "cuenta_id": row[0],
                "codigo": row[1],
                "nombre": row[2]
            })
        cur.close()
        conn.close()
        return cuentas
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/config/gastos")
def get_config_gastos():
    """Obtiene el mapeo actual de cuentas a categorías de gastos."""
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("SELECT categoria, cuenta_codigo, cuenta_nombre FROM cert_config_gastos_cuentas")
        mapeos = cur.fetchall()
        
        config = {}
        for row in mapeos:
            cat = row[0]
            if cat not in config:
                config[cat] = []
            config[cat].append({
                "cuenta_codigo": row[1],
                "cuenta_nombre": row[2]
            })
        
        cur.close()
        conn.close()
        return config
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ═══════════════════════════════════════════════════════
# MAESTROS FINNEGANS (Para Configuración Avanzada)
# ═══════════════════════════════════════════════════════

@app.get("/api/finnegans/subtipos")
def get_finnegans_subtipos():
    try:
        conn = get_aurora()
        cur = conn.cursor()
        cur.execute("SELECT TransaccionSubtipoID, Codigo, Nombre FROM ceesa_faftransaccionsubtipo ORDER BY Nombre")
        data = [{"id": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/finnegans/centros-costo")
def get_finnegans_centros_costo():
    try:
        conn = get_aurora()
        cur = conn.cursor()
        cur.execute("SELECT centrocostoid, codigo, nombre FROM ceesa_bscentrocosto ORDER BY nombre")
        data = [{"id": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/config/centros-costo")
def get_all_config_centros_costo():
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("SELECT sucursal, centro_id, codigo, nombre FROM cert_config_centros_costo")
        data = {}
        for r in cur.fetchall():
            suc = r[0]
            if suc not in data:
                data[suc] = []
            data[suc].append({"id_ref": r[1], "codigo": r[2], "nombre": r[3]})
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/config/centros-costo/{sucursal}")
def get_config_centros_costo(sucursal: str):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("SELECT centro_id, codigo, nombre FROM cert_config_centros_costo WHERE sucursal = %s", (sucursal,))
        data = [{"id_ref": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        return data
    except Exception as e:
        print(f"Error get_config_centros_costo: {e}")
        return []
    finally:
        conn.close()

@app.post("/api/config/centros-costo/{sucursal}")
def save_config_centros_costo(sucursal: str, items: List[dict]):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        try:
            cur.execute("""
                ALTER TABLE cert_config_centros_costo 
                ALTER COLUMN codigo TYPE TEXT, 
                ALTER COLUMN nombre TYPE TEXT, 
                ALTER COLUMN centro_id TYPE TEXT USING centro_id::text
            """)
        except Exception as e:
            conn.rollback()
            print("Alter table failed:", e)
        
        cur.execute("DELETE FROM cert_config_centros_costo WHERE sucursal = %s", (sucursal,))
        for item in items:
            cur.execute("""
                INSERT INTO cert_config_centros_costo (sucursal, centro_id, codigo, nombre)
                VALUES (%s, %s, %s, %s)
            """, (sucursal, item.get("id_ref"), item.get("codigo"), item.get("nombre")))
        conn.commit()
        cur.close()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/api/config/ajustes-excel")
async def upload_ajustes_excel(
    file: UploadFile = File(...), 
    unidad: Optional[str] = Form(None),
    periodo: Optional[str] = Form(None),
    tipo: Optional[str] = Form(None),
    current_user = Depends(get_current_user)
):
    if periodo:
        periodo = normalize_periodo(periodo)
    from datetime import datetime, date
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    conn_aurora = get_aurora()
    cur_a = conn_aurora.cursor()
    cur_a.execute("SELECT TRIM(COALESCE(nombreempresa, '')) FROM ceesa_cee_sucursales WHERE nombreempresa IS NOT NULL")
    valid_sucursales = [r[0] for r in cur_a.fetchall() if r[0]]
    cur_a.execute("SELECT TRIM(COALESCE(Nombre, '')) FROM ceesa_fafempresa WHERE EmpresaIDPadre = '1'")
    valid_sucursales.extend([r[0] for r in cur_a.fetchall() if r[0]])
    cur_a.close()
    conn_aurora.close()
    
    valid_sucursales = list(set(valid_sucursales))
    if "Sede Central" not in valid_sucursales:
        valid_sucursales.append("Sede Central")
    
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    user_email = current_user.get("email", "unknown")
    
    errors = []
    inserted_count = 0
    suggestions = {}
    
    # Leer headers de la fila 1 para determinar dinámicamente si es el nuevo formato de 4 columnas
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    is_new_format = False
    
    if "fecha" in headers and "comprobante" in headers:
        is_new_format = True
        idx_fecha = headers.index("fecha")
        idx_comprobante = headers.index("comprobante")
        idx_concepto = headers.index("concepto") if "concepto" in headers else 2
        idx_importe = headers.index("importe") if "importe" in headers else 3
    else:
        # Fallback si tiene 4 columnas o menos y se proveen los parámetros del form
        if len(headers) <= 4 and unidad and periodo and tipo:
            is_new_format = True
            idx_fecha = 0
            idx_comprobante = 1
            idx_concepto = 2
            idx_importe = 3

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row): continue
        
        if is_new_format:
            fecha_val = row[idx_fecha] if len(row) > idx_fecha else None
            comprobante_val = row[idx_comprobante] if len(row) > idx_comprobante else ""
            concepto_val = row[idx_concepto] if len(row) > idx_concepto else ""
            importe_val = row[idx_importe] if len(row) > idx_importe else None
            
            # Parsear Fecha
            parsed_fecha = None
            if isinstance(fecha_val, datetime):
                parsed_fecha = fecha_val
            elif isinstance(fecha_val, date):
                parsed_fecha = datetime.combine(fecha_val, datetime.min.time())
            elif fecha_val:
                fecha_str = str(fecha_val).strip()
                for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d'):
                    try:
                        parsed_fecha = datetime.strptime(fecha_str, fmt)
                        break
                    except ValueError:
                        continue
            
            if not parsed_fecha:
                errors.append(f"Fila {row_idx}: Fecha '{fecha_val}' no es válida (debe ser DD/MM/AAAA).")
                continue
                
            # Determinar periodo a partir de la fecha (MM/YYYY)
            periodo_val = f"{parsed_fecha.month:02d}/{parsed_fecha.year}"
            
            # Determinar unidad de negocio
            unidad_negocio = str(unidad or "").strip()
            if not unidad_negocio:
                errors.append(f"Fila {row_idx}: No se especificó la Unidad de Negocio.")
                continue
                
            tipo_mov = str(tipo or "").strip().upper()
            if tipo_mov not in ["INGRESO", "COSTO", "GASTO"]:
                errors.append(f"Fila {row_idx}: Tipo de movimiento '{tipo_mov}' inválido.")
                continue
            if tipo_mov == "GASTO":
                tipo_mov = "COSTO"
                
            # Validar y truncar Comprobante
            comprobante = str(comprobante_val or "").strip()
            if len(comprobante) > 15:
                comprobante = comprobante[:15]
            if not comprobante:
                comprobante = "S/N"
                
            # Validar y truncar Concepto
            concepto = str(concepto_val or "").strip()
            if len(concepto) > 30:
                concepto = concepto[:30]
            if not concepto:
                errors.append(f"Fila {row_idx}: El Concepto está vacío.")
                continue
                
            # Validar Importe
            try:
                importe = float(importe_val)
            except (ValueError, TypeError):
                errors.append(f"Fila {row_idx}: Importe '{importe_val}' no es un número.")
                continue
                
            if tipo_mov == "INGRESO":
                categoria = "Ingresos adicionales"
            else:
                categoria = "Egresos adicionales"
            observaciones = comprobante
            fecha_carga = parsed_fecha
        else:
            # Formato viejo
            if len(row) >= 6 and str(row[3] or "").strip().upper() in ["INGRESO", "GASTO", "COSTO"]:
                unidad_negocio = str(row[0] or "").strip()
                periodo_val = str(row[1] or "").strip()
                concepto = str(row[2] or "").strip()
                tipo_mov = str(row[3] or "").strip().upper()
                categoria = str(row[4] or "").strip()
                importe_val = row[5]
                observaciones = str(row[6] or "").strip() if len(row) > 6 else ""
            else:
                unidad_negocio = str(row[0] or "").strip()
                periodo_val = str(row[1] or "").strip()
                concepto = str(row[2] or "").strip()
                tipo_mov = str(row[3] or "").strip().upper()
                categoria = str(row[4] or "").strip()
                importe_val = row[5] if len(row) > 5 else None
                observaciones = str(row[6] or "").strip() if len(row) > 6 else ""
            
            if not unidad_negocio or not periodo_val or not tipo_mov or importe_val is None:
                errors.append(f"Fila {row_idx}: Faltan datos (Unidad, Periodo, Tipo o Importe).")
                continue
                
            if unidad_negocio not in valid_sucursales:
                matches = difflib.get_close_matches(unidad_negocio, valid_sucursales, n=1, cutoff=0.6)
                if matches:
                    suggestions[unidad_negocio] = matches[0]
                    unidad_negocio = matches[0]
                else:
                    errors.append(f"Fila {row_idx}: Unidad '{unidad_negocio}' no encontrada.")
                    continue
                    
            if tipo_mov not in ["INGRESO", "GASTO", "COSTO"]:
                errors.append(f"Fila {row_idx}: Tipo '{tipo_mov}' inválido. Use INGRESO o COSTO.")
                continue
                
            if tipo_mov == "GASTO":
                tipo_mov = "COSTO"
                
            try:
                importe = float(importe_val)
            except (ValueError, TypeError):
                errors.append(f"Fila {row_idx}: Importe '{importe_val}' no es un número.")
                continue
                
            fecha_carga = datetime.now()
            
        cur_supa.execute("""
            INSERT INTO cert_ajustes_excel (unidad_negocio, periodo, concepto, tipo_movimiento, categoria, importe, observaciones, usuario_carga, fecha_carga)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (unidad_negocio, periodo_val, concepto, tipo_mov, categoria, importe, observaciones, user_email, fecha_carga))
        inserted_count += 1
        
    conn_supa.commit()
    cur_supa.close()
    conn_supa.close()
    
    return {
        "status": "ok" if not errors else "partial",
        "inserted": inserted_count,
        "errors": errors,
        "suggestions": suggestions
    }

@app.get("/api/config/ajustes-excel")
def list_ajustes_excel(current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    cur_supa.execute("SELECT id, unidad_negocio, periodo, concepto, tipo_movimiento, categoria, importe, observaciones, fecha_carga, usuario_carga FROM cert_ajustes_excel ORDER BY fecha_carga DESC LIMIT 100")
    rows = cur_supa.fetchall()
    cur_supa.close()
    conn_supa.close()
    
    ajustes = []
    for r in rows:
        db_cat = r[5] or "Ajuste Manual"
        if db_cat == "Ajuste Excel":
            categoria_label = "Ingresos adicionales" if r[4] == "INGRESO" else "Egresos adicionales"
        else:
            categoria_label = db_cat
            
        ajustes.append({
            "id": r[0], "unidad_negocio": r[1], "periodo": r[2], "concepto": r[3],
            "tipo_movimiento": r[4], "categoria": categoria_label, "importe": float(r[6]),
            "observaciones": r[7], "fecha_carga": str(r[8]), "usuario_carga": r[9]
        })
    return ajustes

@app.delete("/api/config/ajustes-excel/{id}")
def delete_ajuste_excel(id: int, current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    cur_supa.execute("DELETE FROM cert_ajustes_excel WHERE id = %s", (id,))
    conn_supa.commit()
    cur_supa.close()
    conn_supa.close()
    return {"status": "ok"}



class BulkDeleteReq(BaseModel):
    ids: List[int]

@app.delete("/api/config/ajustes-excel/bulk")
def delete_ajustes_excel_bulk(req: BulkDeleteReq, current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    if not req.ids:
        return {"status": "ok"}
    
    format_strings = ','.join(['%s'] * len(req.ids))
    cur_supa.execute(f"DELETE FROM cert_ajustes_excel WHERE id IN ({format_strings})", tuple(req.ids))
    conn_supa.commit()
    cur_supa.close()
    conn_supa.close()
    return {"status": "ok"}


class EquiposAsignadosReq(BaseModel):
    sucursal: str
    equipo_codigo: str
    equipo_nombre: str
    fecha_desde: str
    fecha_hasta: Optional[str] = None


@app.get("/api/equipos/maestro")
def get_equipos_maestro(current_user = Depends(get_current_user)):
    import urllib.request
    import json
    
    url = "https://naxjzquhdzyoxtjataaw.supabase.co/rest/v1/equipos?select=codigo,nombre"
    apikey = "sb_publishable_K3r5ogAi25D2ffI6-9cWLg_VJddbdxF"
    req = urllib.request.Request(
        url,
        headers={
            "apikey": apikey,
            "Authorization": f"Bearer {apikey}"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            data.sort(key=lambda x: str(x.get('nombre') or '').lower())
            return data
    except Exception as e:
        print("Error fetching equipos maestro:", e)
        raise HTTPException(status_code=500, detail=f"Error al obtener catálogo de equipos: {str(e)}")


@app.get("/api/config/equipos-asignados")
def list_equipos_asignados(current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    cur_supa.execute("""
        SELECT id, sucursal, equipo_codigo, equipo_nombre, fecha_desde, fecha_hasta, created_at, usuario_carga, fecha_carga 
        FROM cert_config_equipos_asignados 
        ORDER BY sucursal, fecha_desde DESC
    """)
    rows = cur_supa.fetchall()
    cur_supa.close()
    conn_supa.close()
    
    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "sucursal": r[1],
            "equipo_codigo": r[2],
            "equipo_nombre": r[3],
            "fecha_desde": str(r[4]) if r[4] else "",
            "fecha_hasta": str(r[5]) if r[5] else "",
            "created_at": str(r[6]),
            "usuario_carga": r[7],
            "fecha_carga": str(r[8])
        })
    return result


@app.post("/api/config/equipos-asignados")
def create_equipo_asignado(req: EquiposAsignadosReq, current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    
    try:
        fd = datetime.strptime(req.fecha_desde, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="fecha_desde inválida. Formato esperado: YYYY-MM-DD")
        
    fh = None
    if req.fecha_hasta:
        try:
            fh = datetime.strptime(req.fecha_hasta, "%Y-%m-%d").date()
            if fh < fd:
                raise HTTPException(status_code=400, detail="fecha_hasta no puede ser anterior a fecha_desde")
        except HTTPException as he:
            raise he
        except Exception:
            raise HTTPException(status_code=400, detail="fecha_hasta inválida. Formato esperado: YYYY-MM-DD")

    user_email = current_user.get("email") or "usuario"
    cur_supa.execute("""
        INSERT INTO cert_config_equipos_asignados (sucursal, equipo_codigo, equipo_nombre, fecha_desde, fecha_hasta, usuario_carga)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (req.sucursal, req.equipo_codigo, req.equipo_nombre, fd, fh, user_email))
    conn_supa.commit()
    cur_supa.close()
    conn_supa.close()
    
    log_action(user_email, "ASIGNAR_EQUIPO", f"Asignó {req.equipo_nombre} ({req.equipo_codigo}) a {req.sucursal} desde {req.fecha_desde}")
    return {"status": "ok"}


@app.delete("/api/config/equipos-asignados/{id}")
def delete_equipo_asignado(id: int, current_user = Depends(get_current_user)):
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    
    cur_supa.execute("SELECT sucursal, equipo_nombre, equipo_codigo FROM cert_config_equipos_asignados WHERE id = %s", (id,))
    row = cur_supa.fetchone()
    if not row:
        cur_supa.close()
        conn_supa.close()
        raise HTTPException(status_code=404, detail="Configuración no encontrada")
        
    cur_supa.execute("DELETE FROM cert_config_equipos_asignados WHERE id = %s", (id,))
    conn_supa.commit()
    cur_supa.close()
    conn_supa.close()
    
    user_email = current_user.get("email") or "usuario"
    log_action(user_email, "DESASIGNAR_EQUIPO", f"Eliminó asignación de {row[1]} ({row[2]}) de sucursal {row[0]}")
    return {"status": "ok"}


@app.get("/api/finnegans/empresas")
def get_finnegans_empresas():
    try:
        conn = get_aurora()
        cur = conn.cursor()
        cur.execute("SELECT EmpresaID, Codigo, Nombre FROM ceesa_fafempresa WHERE EmpresaIDPadre = '1' ORDER BY Nombre")
        data = [{"id": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/finnegans/categorias-asiento")
def get_finnegans_categorias_asiento():
    try:
        conn = get_aurora()
        cur = conn.cursor()
        cur.execute("SELECT TransaccionCategoriaID, Codigo, Nombre FROM ceesa_faftransaccioncategoria ORDER BY Nombre")
        data = [{"id": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════
# CRUD CONFIGURACIÓN AVANZADA
# ═══════════════════════════════════════════════════════



class ConfigItem(BaseModel):
    id_ref: str
    codigo: str
    nombre: str

@app.get("/api/config/avanzada/{tipo}")
def get_all_config_avanzada(tipo: str):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        
        if tipo == "ingresos-comprobantes":
            cur.execute("SELECT sucursal, subtipo_id, codigo, nombre FROM cert_config_ingresos_comprobantes")
        elif tipo == "gastos-asientos":
            cur.execute("SELECT sucursal, tipo_asiento_id, codigo, nombre FROM cert_config_gastos_asientos")
        elif tipo == "gastos-compras":
            cur.execute("SELECT sucursal, subtipo_id, codigo, nombre FROM cert_config_gastos_compras")
        else:
            raise HTTPException(status_code=400, detail="Tipo de configuración inválido")
            
        data = {}
        for r in cur.fetchall():
            suc = r[0]
            if suc not in data:
                data[suc] = []
            data[suc].append({"id_ref": r[1], "codigo": r[2], "nombre": r[3]})
            
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/config/avanzada/{tipo}/{sucursal}")
def get_config_avanzada(tipo: str, sucursal: str):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        
        if tipo == "ingresos-comprobantes":
            cur.execute("SELECT subtipo_id, codigo, nombre FROM cert_config_ingresos_comprobantes WHERE sucursal = %s", (sucursal,))
        elif tipo == "gastos-asientos":
            cur.execute("SELECT tipo_asiento_id, codigo, nombre FROM cert_config_gastos_asientos WHERE sucursal = %s", (sucursal,))
        elif tipo == "gastos-compras":
            cur.execute("SELECT subtipo_id, codigo, nombre FROM cert_config_gastos_compras WHERE sucursal = %s", (sucursal,))
        else:
            raise HTTPException(status_code=400, detail="Tipo de configuración inválido")
            
        data = [{"id_ref": r[0], "codigo": r[1], "nombre": r[2]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/config/avanzada/{tipo}/{sucursal}")
def save_config_avanzada(tipo: str, sucursal: str, items: List[ConfigItem]):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        
        if tipo == "ingresos-comprobantes":
            cur.execute("DELETE FROM cert_config_ingresos_comprobantes WHERE sucursal = %s", (sucursal,))
            for item in items:
                cur.execute("INSERT INTO cert_config_ingresos_comprobantes (sucursal, subtipo_id, codigo, nombre) VALUES (%s, %s, %s, %s)",
                            (sucursal, item.id_ref, item.codigo, item.nombre))
        elif tipo == "gastos-asientos":
            cur.execute("DELETE FROM cert_config_gastos_asientos WHERE sucursal = %s", (sucursal,))
            for item in items:
                cur.execute("INSERT INTO cert_config_gastos_asientos (sucursal, tipo_asiento_id, codigo, nombre) VALUES (%s, %s, %s, %s)",
                            (sucursal, item.id_ref, item.codigo, item.nombre))
        elif tipo == "gastos-compras":
            cur.execute("DELETE FROM cert_config_gastos_compras WHERE sucursal = %s", (sucursal,))
            for item in items:
                cur.execute("INSERT INTO cert_config_gastos_compras (sucursal, subtipo_id, codigo, nombre) VALUES (%s, %s, %s, %s)",
                            (sucursal, item.id_ref, item.codigo, item.nombre))
        else:
            raise HTTPException(status_code=400, detail="Tipo de configuración inválido")
            
        conn.commit()
        cur.close()
        conn.close()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





class ExcelItem(BaseModel):
    unidad_negocio: str
    fecha: str
    concepto: str
    tipo: str
    importe: float

@app.post("/api/excel/upload")
async def upload_excel(file: UploadFile = File(...)):
    import openpyxl
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        sheet = wb.active
        
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        # Validar columnas
        expected_cols = ["Unidad de Negocio", "Fecha", "Concepto", "Tipo", "Importe"]
        for col in expected_cols:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Falta la columna '{col}' en el Excel. Encontradas: {headers}")
                
        col_idx = {col: headers.index(col) for col in expected_cols}
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        # Filtrar filas vacías
        valid_rows = [r for r in rows if any(x is not None for x in r)]
                
        # Insertar cabecera
        conn = get_supabase()
        cur = conn.cursor()
        
        cur.execute(
            "INSERT INTO cert_excel_uploads (filename, estado, total_registros) VALUES (%s, %s, %s) RETURNING id",
            (file.filename, 'PROCESADO', len(valid_rows))
        )
        upload_id = cur.fetchone()[0]
        
        # Insertar items
        for row in valid_rows:
            raw_importe = row[col_idx['Importe']]
            importe = 0
            if raw_importe is not None:
                try:
                    importe = float(str(raw_importe).replace(',', '.'))
                except:
                    importe = 0
                    
            raw_fecha = row[col_idx['Fecha']]
            fecha_str = datetime.now().strftime('%Y-%m-%d')
            if raw_fecha is not None:
                if isinstance(raw_fecha, datetime):
                    fecha_str = raw_fecha.strftime('%Y-%m-%d')
                else:
                    try:
                        # Attempt to parse common string formats if it's a string
                        from dateutil import parser
                        fecha_str = parser.parse(str(raw_fecha)).strftime('%Y-%m-%d')
                    except:
                        pass
                        
            cur.execute("""
                INSERT INTO cert_excel_items (upload_id, unidad_negocio, fecha, concepto, tipo, importe)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                upload_id,
                str(row[col_idx['Unidad de Negocio']]) if row[col_idx['Unidad de Negocio']] is not None else 'DESCONOCIDO',
                fecha_str,
                str(row[col_idx['Concepto']]) if row[col_idx['Concepto']] is not None else '',
                str(row[col_idx['Tipo']]).upper() if row[col_idx['Tipo']] is not None else 'INGRESO',
                importe
            ))
            
        conn.commit()
        cur.close()
        conn.close()
        
        return {"status": "ok", "id": upload_id, "registros": len(df)}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/uploads")
def get_excel_uploads():
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("SELECT id, filename, uploaded_at, estado, total_registros FROM cert_excel_uploads ORDER BY uploaded_at DESC")
        data = [{"id": r[0], "filename": r[1], "uploaded_at": str(r[2]), "estado": r[3], "total_registros": r[4]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/excel/uploads/{id}")
def delete_excel_upload(id: int):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("DELETE FROM cert_excel_uploads WHERE id = %s", (id,))
        conn.commit()
        cur.close()
        conn.close()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
class GastoConfigUpdate(BaseModel):
    categoria: str
    cuentas: List[dict]

@app.post("/api/config/gastos")
def update_config_gastos(data: GastoConfigUpdate):
    """Actualiza las cuentas asociadas a una categoría de gasto específica."""
    try:
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("DELETE FROM cert_config_gastos_cuentas WHERE categoria = %s", (data.categoria,))
        
        for c in data.cuentas:
            cur.execute("""
                INSERT INTO cert_config_gastos_cuentas (categoria, cuenta_codigo, cuenta_nombre)
                VALUES (%s, %s, %s)
            """, (data.categoria, c.get("cuenta_codigo"), c.get("cuenta_nombre")))
        
        conn.commit()
        cur.close()
        conn.close()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/asientos")
def get_asientos(
    empresa: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    try:
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        # Obtener los subtipos de asiento configurados para esta Sucursal
        cur_supa.execute("SELECT tipo_asiento_id FROM cert_config_gastos_asientos WHERE sucursal = %s", (empresa,))
        tipos_rows = cur_supa.fetchall()
        tipos_asientos = [r[0] for r in tipos_rows]
        
        # Obtener centros de costo
        cur_supa.execute("SELECT centro_id FROM cert_config_centros_costo WHERE sucursal = %s", (empresa,))
        centros = [r[0] for r in cur_supa.fetchall()]
        
        # Obtener empresa padre
        conn_a_tmp = get_aurora()
        cur_a_tmp = conn_a_tmp.cursor()
        cur_a_tmp.execute("""
            SELECT MAX(TRIM(COALESCE(nombreempresapadre, '')))
            FROM ceesa_cee_sucursales
            WHERE TRIM(COALESCE(nombreempresa, '')) = %s
        """, (empresa,))
        padre_row = cur_a_tmp.fetchone()
        empresa_padre = padre_row[0] if padre_row and padre_row[0] else empresa
        cur_a_tmp.close()
        conn_a_tmp.close()

        cur_supa.close()
        conn_supa.close()
        
        if not tipos_asientos or not centros:
            return []
            
        empresas_validas = list(set([empresa, empresa_padre]))

        conn = get_aurora()
        cur = conn.cursor()

        sql = """
        SELECT 
            BSAsientoItem.Fecha AS fecha,
            BSCuenta.Codigo AS cuenta_codigo,
            BSCuenta.Nombre AS cuenta_nombre,
            BSAsientoItem.Descripcion AS descripcion,
            (COALESCE(BSAsientoItem.ImporteMonPrincipal, 0) * BSAsientoItem.DebeHaber) AS importe
        FROM BSAsientoItem
        INNER JOIN BSCuenta ON BSAsientoItem.CuentaID = BSCuenta.CuentaID
        INNER JOIN BSTransaccion ON BSAsientoItem.TransaccionID = BSTransaccion.TransaccionID
        INNER JOIN FAFEmpresa ON BSTransaccion.EmpresaID = FAFEmpresa.EmpresaID
        LEFT JOIN BSTransaccionDimension 
            ON BSAsientoItem.AsientoItemID = BSTransaccionDimension.AsientoItemID 
            AND BSTransaccion.TransaccionID = BSTransaccionDimension.TransaccionID
            AND BSTransaccionDimension.DimensionID = '999999'
        WHERE BSTransaccion.TransaccionSubtipoID IN %s
        """
        
        params = [tuple(tipos_asientos)]
        
        empresas_str = ",".join(f"'{e}'" for e in empresas_validas)
        ccs_str = ",".join(f"'{c}'" for c in centros)
        sql += f" AND FAFEmpresa.Nombre IN ({empresas_str}) AND BSTransaccionDimension.RegistroID IN ({ccs_str})"
        if fecha_desde:
            sql += " AND BSAsientoItem.Fecha >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND BSAsientoItem.Fecha <= %s"
            params.append(fecha_hasta)
            
        sql += " ORDER BY BSAsientoItem.Fecha DESC LIMIT 1000"
        
        cur.execute(sql, tuple(params))
        cols = [desc[0].lower() for desc in cur.description]
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        return [dict(zip(cols, row)) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/rrhh")
def get_rrhh(
    empresa: Optional[str] = None,
    periodo: Optional[str] = None
):
    try:
        periodo = normalize_periodo(periodo)
        cerrado = check_informe_cerrado(empresa, periodo, 'rrhh')
        if cerrado is not None: return cerrado
        empty_response = {
            "legajos": [],
            "totales": {
                "remunerativo": 0.0,
                "no_remunerativo": 0.0,
                "contribuciones": 0.0,
                "retenciones": 0.0,
                "sac_prorrateado": 0.0,
                "costo_empresa": 0.0,
                "neto": 0.0
            }
        }
        
        if not empresa or not periodo:
            return empty_response
            
        # El periodo viene como 'YYYY-MM', para RRHH lo pasamos a 'YYYYMM'
        periodo_str = periodo.replace("-", "")
        year_str, month_str = periodo.split('-')
        is_aguinaldo = month_str in ('06', '12')
        semestre = f"{year_str}-S1" if int(month_str) <= 6 else f"{year_str}-S2"

        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        # Obtener los centros de costo configurados para esta Sucursal
        cur_supa.execute("SELECT nombre FROM cert_config_centros_costo WHERE sucursal = %s", (empresa,))
        centros_rows = cur_supa.fetchall()
        
        centros_costo = [r[0] for r in centros_rows]
        if not centros_costo:
            cur_supa.close()
            conn_supa.close()
            return empty_response

        # Obtener provisiones de SAC de Supabase para este semestre (hasta el mes anterior)
        # Si es Junio, restamos lo acumulado de Ene a May. Si es Dic, de Jul a Nov.
        provisiones_por_legajo = {}
        # NOTA: Simplificaremos obteniendo la suma total de provisiones guardada para la unidad de negocio
        # en este semestre, o mejor dicho, la provisión fue guardada a nivel de "Unidad de Negocio" mes a mes.
        
        cur_supa.execute("""
            SELECT SUM(monto_provision) 
            FROM cert_rrhh_provisiones_sac 
            WHERE unidad_negocio = %s AND semestre = %s AND periodo != %s
        """, (empresa, semestre, periodo))
        row_prov = cur_supa.fetchone()
        suma_provisiones_semestre = float(row_prov[0] or 0) if row_prov else 0.0

        cur_supa.close()
        conn_supa.close()

        conn = get_aurora()
        cur = conn.cursor()

        # Obtener la empresa padre para filtrar la liquidacion
        cur.execute("""
            SELECT MAX(TRIM(COALESCE(nombreempresapadre, '')))
            FROM ceesa_cee_sucursales
            WHERE TRIM(COALESCE(nombreempresa, '')) = %s
        """, (empresa,))
        padre_row = cur.fetchone()
        empresa_padre = padre_row[0] if padre_row and padre_row[0] else empresa

        # Agrupamos por legajo, apellidonombre, centrocosto, tipoconcepto y nombreconcepto
        sql = """
        WITH distinct_rows AS (
            SELECT DISTINCT
                transaccionid,
                legajo,
                apellidonombre,
                centrocosto,
                tipoconcepto,
                codigoconcepto,
                nombreconcepto,
                CAST(REPLACE(importe, ',', '.') AS NUMERIC) as importe
            FROM ceesa_cee_liquidaciones_de_sueldos_
            WHERE periodo = %s
              AND centrocosto IN %s
              AND TRIM(COALESCE(empresa, '')) = %s
        )
        SELECT
            legajo,
            apellidonombre,
            centrocosto,
            tipoconcepto,
            nombreconcepto,
            SUM(importe) as importe
        FROM distinct_rows
        GROUP BY
            legajo, apellidonombre, centrocosto, tipoconcepto, nombreconcepto
        """
        
        cur.execute(sql, (periodo_str, tuple(centros_costo), empresa_padre))
        cols = [desc[0].lower() for desc in cur.description]
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        legajos_map = {}
        totales = {
            "remunerativo": 0.0,
            "no_remunerativo": 0.0,
            "contribuciones": 0.0,
            "retenciones": 0.0,
            "sac_prorrateado": 0.0,
            "sac_real": 0.0
        }
        
        for row in rows:
            record = dict(zip(cols, row))
            leg = record['legajo']
            if leg not in legajos_map:
                legajos_map[leg] = {
                    "legajo": leg,
                    "apellidonombre": record['apellidonombre'],
                    "centrocosto": record['centrocosto'],
                    "remunerativo": 0.0,
                    "no_remunerativo": 0.0,
                    "contribuciones": 0.0,
                    "retenciones": 0.0,
                    "sac_prorrateado": 0.0,
                    "sac_real": 0.0
                }
                
            tipo = str(record['tipoconcepto']).strip()
            nombre = str(record['nombreconcepto']).strip().upper()
            imp = float(record['importe'])
            
            es_sac = ('SAC' in nombre or 'AGUINALDO' in nombre)
            
            if es_sac:
                legajos_map[leg]['sac_real'] += imp
                totales['sac_real'] += imp
            elif tipo in ('Remunerativo', 'Remunerativo Variable'):
                legajos_map[leg]['remunerativo'] += imp
                totales['remunerativo'] += imp
            elif tipo == 'No Remunerativo':
                legajos_map[leg]['no_remunerativo'] += imp
                totales['no_remunerativo'] += imp
            elif tipo == 'Contribución Patronal' or tipo == 'Contribucion Patronal' or tipo == 'Contribucin Patronal' or 'Contrib' in tipo:
                legajos_map[leg]['contribuciones'] += imp
                totales['contribuciones'] += imp
            elif tipo == 'Retención' or tipo == 'Retencion' or tipo == 'Retencin' or 'Retenc' in tipo:
                legajos_map[leg]['retenciones'] += imp
                totales['retenciones'] += imp

        # Computar Costo Empresa y Neto
        resultado_legajos = []
        total_sac_prorrateado_mes = 0.0

        for leg in legajos_map.values():
            rem = leg['remunerativo']
            no_rem = leg['no_remunerativo']
            cont = leg['contribuciones']
            ret = abs(leg['retenciones'])
            sac_real = leg['sac_real']
            
            if is_aguinaldo:
                # En meses de SAC (Jun/Dic), no hay prorrateo para costo, el costo incluye el sac_real
                # PERO al SAC Real hay que descontarle las provisiones acumuladas de ese legajo!
                # Como guardamos la provision por Unidad de Negocio y no por legajo (porque era más sencillo en el plan original),
                # lo aplicaremos como un ajuste global en los totales, o lo dividiremos proporcionalmente.
                # Para simplificar, guardamos el SAC prorrateado en 'sac_prorrateado' como el SAC Real de este mes 
                # (el ajuste global lo haremos despues sobre los totales, o dividiremos suma_provisiones_semestre entre los legajos).
                # Mejor aún: el informe muestra los legajos. El prorrateo de SAC de meses anteriores no afecta el 'Neto' que cobra el empleado.
                # Afecta el COSTO EMPRESA.
                # Costo Empresa Mes = Remunerativo + No Remunerativo + Contribuciones + SAC Real - Provisiones Aplicadas
                leg['sac_prorrateado'] = 0.0
                leg['retenciones'] = ret
                leg['costo_empresa'] = rem + no_rem + cont + sac_real
                leg['neto'] = rem + no_rem + sac_real - ret
            else:
                # Meses no SAC: se provisiona 8.33% de remunerativo + contribuciones
                sac_prorr = (rem / 12.0) + (cont / 12.0)
                leg['sac_prorrateado'] = sac_prorr
                total_sac_prorrateado_mes += sac_prorr
                
                leg['retenciones'] = ret
                leg['costo_empresa'] = rem + no_rem + cont + sac_prorr
                leg['neto'] = rem + no_rem - ret
                
            resultado_legajos.append(leg)

        totales['retenciones'] = abs(totales['retenciones'])
        
        if is_aguinaldo:
            # Descontar la provisión acumulada del SAC real en los totales
            sac_ajustado = totales['sac_real'] - suma_provisiones_semestre
            totales['sac_prorrateado'] = sac_ajustado  # Lo mostramos en la columna sac_prorrateado para el dashboard
            totales['costo_empresa'] = totales['remunerativo'] + totales['no_remunerativo'] + totales['contribuciones'] + sac_ajustado
            # El neto que cobra el empleado sí incluye el SAC Real completo
            totales['neto'] = totales['remunerativo'] + totales['no_remunerativo'] + totales['sac_real'] - totales['retenciones']
        else:
            totales['sac_prorrateado'] = total_sac_prorrateado_mes
            totales['costo_empresa'] = totales['remunerativo'] + totales['no_remunerativo'] + totales['contribuciones'] + totales['sac_prorrateado']
            totales['neto'] = totales['remunerativo'] + totales['no_remunerativo'] - totales['retenciones']
            
            # Guardar provisión mensual en Supabase
            conn_s = get_supabase()
            cur_s = conn_s.cursor()
            cur_s.execute("""
                INSERT INTO cert_rrhh_provisiones_sac (unidad_negocio, periodo, semestre, monto_provision)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (unidad_negocio, periodo) 
                DO UPDATE SET monto_provision = EXCLUDED.monto_provision
            """, (empresa, periodo, semestre, total_sac_prorrateado_mes))
            conn_s.commit()
            cur_s.close()
            conn_s.close()

        return {
            "totales": totales,
            "legajos": resultado_legajos
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
def get_gastos(
    empresa: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None
):
    try:
        # Asumimos que fecha_desde y fecha_hasta son el mismo periodo para el cierre
        cerrado = check_informe_cerrado(empresa, fecha_desde, 'gastos')
        if cerrado is not None: return cerrado
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        # Obtener centros de costo
        cur_supa.execute("SELECT centro_id FROM cert_config_centros_costo WHERE sucursal = %s", (empresa,))
        centros = [r[0] for r in cur_supa.fetchall()]
        
        # Obtener empresa padre
        conn_a_tmp = get_aurora()
        cur_a_tmp = conn_a_tmp.cursor()
        cur_a_tmp.execute("""
            SELECT MAX(TRIM(COALESCE(nombreempresapadre, '')))
            FROM ceesa_cee_sucursales
            WHERE TRIM(COALESCE(nombreempresa, '')) = %s
        """, (empresa,))
        padre_row = cur_a_tmp.fetchone()
        empresa_padre = padre_row[0] if padre_row and padre_row[0] else empresa
        cur_a_tmp.close()
        conn_a_tmp.close()

        cur_supa.execute("SELECT categoria, cuenta_codigo FROM cert_config_gastos_cuentas")
        config_rows = cur_supa.fetchall()
        cur_supa.close()
        conn_supa.close()
        
        if not centros:
            return []
            
        cat_map = {}
        for r in config_rows:
            cat_map[r[1]] = r[0]

        empresas_validas = list(set([empresa, empresa_padre]))

        conn = get_aurora()
        cur = conn.cursor()

        sql = """
        SELECT
            FAFEmpresa.Nombre AS Sucursal,
            BSCentroCosto.Nombre AS CentroCosto,
            BSCuenta.Codigo AS CuentaCodigo,
            BSCuenta.Nombre AS Cuenta,
            EXTRACT(YEAR FROM BSAsientoItem.Fecha) AS Anio,
            EXTRACT(MONTH FROM BSAsientoItem.Fecha) AS Mes,
            SUM(COALESCE(BSTransaccionDimension.ImporteMonPrincipal, BSAsientoItem.ImporteMonPrincipal) * BSAsientoItem.DebeHaber) AS Importe
        FROM BSAsientoItem
        INNER JOIN BSCuenta ON BSAsientoItem.CuentaID = BSCuenta.CuentaID
        LEFT JOIN BSTransaccionDimension ON BSAsientoItem.AsientoItemID = BSTransaccionDimension.AsientoItemID AND BSTransaccionDimension.DimensionID = '999999'
        INNER JOIN BSTransaccion ON COALESCE(BSTransaccionDimension.TransaccionID, BSAsientoItem.TransaccionID) = BSTransaccion.TransaccionID
        INNER JOIN FAFTransaccionSubtipo ON BSTransaccion.TransaccionSubtipoID = FAFTransaccionSubtipo.TransaccionSubtipoID
        INNER JOIN FAFTransaccionCategoria ON FAFTransaccionSubtipo.TransaccionCategoriaID = FAFTransaccionCategoria.TransaccionCategoriaID
        INNER JOIN FAFEmpresa ON BSTransaccion.EmpresaID = FAFEmpresa.EmpresaID
        LEFT JOIN BSCentroCosto ON BSTransaccionDimension.RegistroID = BSCentroCosto.CentroCostoID
        WHERE BSCuenta.ImpactaResultados = 1
        """
        
        empresas_str = ",".join(f"'{e}'" for e in empresas_validas)
        ccs_str = ",".join(f"'{c}'" for c in centros)
        sql += f" AND FAFEmpresa.Nombre IN ({empresas_str}) AND BSTransaccionDimension.RegistroID IN ({ccs_str})"
        
        params = []
        if fecha_desde:
            sql += " AND BSAsientoItem.Fecha >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND BSAsientoItem.Fecha <= %s"
            params.append(fecha_hasta)
            
        sql += """
        GROUP BY 
            FAFEmpresa.Nombre,
            BSCentroCosto.Nombre,
            BSCuenta.Codigo,
            BSCuenta.Nombre,
            EXTRACT(YEAR FROM BSAsientoItem.Fecha),
            EXTRACT(MONTH FROM BSAsientoItem.Fecha)
        """
        
        cur.execute(sql, tuple(params))
        results = cur.fetchall()
        
        gastos = []
        for row in results:
            sucursal = row[0]
            if empresa and sucursal != empresa and empresa != "Todas":
                continue 
                
            cc = row[1]
            cod = row[2]
            cta = row[3]
            anio = int(row[4])
            mes = int(row[5])
            importe = float(row[6] or 0)
            
            if importe == 0:
                continue
                
            categoria = cat_map.get(cod)
            if not categoria:
                n = (cta or "").upper()
                if "SUELDO" in n or "SALARIO" in n or "REMUNERACION" in n:
                    categoria = "Sueldos"
                elif "COMBUSTIBLE" in n or "GASOIL" in n or "NAFTA" in n:
                    categoria = "Combustible"
                elif "INSUMO" in n or "MATERIAL" in n:
                    categoria = "Insumos"
                elif "COMEDOR" in n or "VIANDA" in n:
                    categoria = "Comedor"
                elif "ALQUILER" in n or "LOCACION" in n or "OFICINA" in n:
                    categoria = "Oficina Central"
                else:
                    categoria = "Otros Gastos"
                    
            gastos.append({
                "sucursal": sucursal,
                "centro_costo": cc,
                "categoria": categoria,
                "cuenta_codigo": cod,
                "cuenta_nombre": cta,
                "periodo": f"{mes:02d}/{anio}",
                "importe": importe
            })
            
        cur.close()
        conn.close()
        return {"gastos": gastos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/login")
def login(req: LoginRequest):
    auto_setup_db()  # Asegurar tablas existen
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, email, nombre, password, rol, activo FROM cert_usuarios WHERE email = %s", (req.email,))
        row = cur.fetchone()
        cur.close()
        
        if not row:
            raise HTTPException(status_code=401, detail="Credenciales inválidas")
        
        user_id, email, nombre, pwd_hash, rol, activo = row
        
        if not activo:
            raise HTTPException(status_code=401, detail="Usuario desactivado")
        
        if not verify_password(req.password, pwd_hash):
            log_action(req.email, "LOGIN_FAILED", "Intento de login con contraseña incorrecta")
            raise HTTPException(status_code=401, detail="Credenciales inválidas")
        
        token = create_token(user_id, email, rol, nombre)
        log_action(email, "LOGIN", "Inicio de sesión exitoso")
        return {"token": token, "user": {"id": user_id, "email": email, "nombre": nombre, "rol": rol}}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)[:200]}")
    finally:
        conn.close()

@app.get("/api/me")
def get_me(user=Depends(get_current_user)):
    return {"user": user}

@app.get("/api/mis-unidades")
def get_mis_unidades(user=Depends(get_current_user)):
    user_id = user.get("id") or user.get("sub")
    rol = user.get("rol")
    
    # Si es admin, puede ver todas las sucursales (leemos de Aurora o simplemente devolvemos las de Aurora)
    if rol == "admin":
        try:
            conn = get_aurora()
            cur = conn.cursor()
            cur.execute("""
                SELECT TRIM(COALESCE(nombreempresa, '')) as sucursal
                FROM ceesa_cee_sucursales
                WHERE nombreempresa IS NOT NULL AND TRIM(COALESCE(nombreempresa, '')) != ''
                GROUP BY TRIM(COALESCE(nombreempresa, ''))
                ORDER BY sucursal
            """)
            lista = [{"id": r[0], "nombre": r[0]} for r in cur.fetchall() if r[0]]
            cur.close()
            conn.close()
            # Añadir 'General' o 'Todas' si hace falta, aunque el frontend tiene un select con option ""
            return lista
        except Exception as e:
            print("Error get_mis_unidades (admin):", e)
            return []
            
    # Si es consulta o responsable_un, solo ve las asignadas
    conn_supa = get_supabase()
    try:
        cur = conn_supa.cursor()
        cur.execute("SELECT unidad_negocio FROM cert_usuarios_unidades WHERE usuario_id = %s ORDER BY unidad_negocio", (user_id,))
        rows = cur.fetchall()
        cur.close()
        return [{"id": r[0], "nombre": r[0]} for r in rows]
    except Exception as e:
        print("Error get_mis_unidades:", e)
        return []
    finally:
        conn_supa.close()

# ═══════════════════════════════════════════════════════
# CRUD USUARIOS (Solo Admin)
# ═══════════════════════════════════════════════════════

@app.get("/api/usuarios")
def list_usuarios(admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.email, u.nombre, u.rol, u.telegram_chat_id, u.activo, u.created_at,
                   (SELECT COUNT(*) FROM cert_usuarios_unidades cu WHERE cu.usuario_id = u.id) as sucursales_asignadas
            FROM cert_usuarios u
            ORDER BY u.id
        """)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()
        return {"data": rows}
    finally:
        conn.close()

@app.post("/api/usuarios")
def create_usuario(body: UsuarioCreate, admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        pwd_hash = hash_password(body.password)
        cur.execute("""
            INSERT INTO cert_usuarios (email, nombre, password, rol, telegram_chat_id, activo)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (body.email, body.nombre, pwd_hash, body.rol, body.telegram_chat_id, body.activo))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        log_action(admin["email"], "NUEVO_USUARIO", f"Creó usuario {body.email} (Rol: {body.rol})")
        return {"ok": True, "id": new_id}
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        raise HTTPException(status_code=400, detail="Email ya registrado")
    finally:
        conn.close()

@app.put("/api/usuarios/{user_id}")
def update_usuario(user_id: int, body: UsuarioUpdate, admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        updates = []
        params = []
        
        if body.nombre is not None:
            updates.append("nombre = %s"); params.append(body.nombre)
        if body.email is not None:
            updates.append("email = %s"); params.append(body.email)
        if body.password and body.password.strip():
            updates.append("password = %s"); params.append(hash_password(body.password))
        if body.rol is not None:
            updates.append("rol = %s"); params.append(body.rol)
        if body.telegram_chat_id is not None:
            updates.append("telegram_chat_id = %s"); params.append(body.telegram_chat_id)
        if body.activo is not None:
            updates.append("activo = %s"); params.append(body.activo)
        
        if not updates:
            return {"ok": True}
        
        params.append(user_id)
        cur.execute(f"UPDATE cert_usuarios SET {', '.join(updates)} WHERE id = %s", params)
        conn.commit()
        cur.close()
        log_action(admin["email"], "ACTUALIZAR_USUARIO", f"Actualizó usuario ID {user_id}")
        return {"ok": True}
    finally:
        conn.close()

@app.delete("/api/usuarios/{user_id}")
def delete_usuario(user_id: int, admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM cert_usuarios WHERE id = %s", (user_id,))
        conn.commit()
        cur.close()
        log_action(admin["email"], "ELIMINAR_USUARIO", f"Eliminó usuario ID {user_id}")
        return {"ok": True}
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════
# GESTIÓN DE UNIDADES DE NEGOCIO POR USUARIO
# ═══════════════════════════════════════════════════════

@app.get("/api/usuarios/{user_id}/unidades")
def get_user_unidades(user_id: int, user=Depends(get_current_user)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT unidad_negocio 
            FROM cert_usuarios_unidades WHERE usuario_id = %s ORDER BY unidad_negocio
        """, (user_id,))
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()
        return {"data": rows}
    except Exception as e:
        print(f"Error en get_user_unidades: {e}")
        return {"data": []}
    finally:
        conn.close()

@app.put("/api/usuarios/{user_id}/unidades")
def update_user_unidades(user_id: int, body: List[UnidadAsignacion], admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        # Borrar asignaciones previas
        cur.execute("DELETE FROM cert_usuarios_unidades WHERE usuario_id = %s", (user_id,))
        # Insertar nuevas asegurando que no haya duplicados si el frontend los envía
        vistos = set()
        for u in body:
            un_limpio = u.unidad_negocio.strip()
            if un_limpio not in vistos:
                cur.execute("""
                    INSERT INTO cert_usuarios_unidades (usuario_id, unidad_negocio)
                    VALUES (%s, %s)
                """, (user_id, un_limpio))
                vistos.add(un_limpio)
        conn.commit()
        cur.close()
        log_action(admin["email"], "ACTUALIZAR_SUCURSALES", f"Actualizó sucursales para usuario ID {user_id}")
        return {"ok": True}
    finally:
        conn.close()

@app.get("/api/audit-logs")
def get_audit_logs(admin=Depends(require_admin)):
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, usuario_email, accion, detalles, fecha FROM cert_audit_log ORDER BY fecha DESC LIMIT 500")
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()
        return {"data": rows}
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════
# CRON & BOT
# ═══════════════════════════════════════════════════════

@app.get("/api/cron/notificar")
def cron_notificar():
    """Endpoint llamado periódicamente por Vercel Cron para notificar nuevos certificados."""
    conn_supa = get_supabase()
    conn_aurora = None
    try:
        cur_supa = conn_supa.cursor()
        cur_supa.execute("SELECT comprobante, estado_doc FROM cert_notificaciones_log WHERE tipo = 'cron_telegram'")
        ya_notificados = {row[0]: row[1] for row in cur_supa.fetchall()}
        
        conn_aurora = get_aurora()
        cur_aurora = conn_aurora.cursor()
        # Obtener los últimos 200 comprobantes con un SELECT * para obtener esquema dinámico
        cur_aurora.execute("""
            SELECT *
            FROM ceesa_cee_certificados_ventas_internas
            WHERE numerodocumento IS NOT NULL AND numerodocumento != '' AND numerodocumento != 'NULL'
            ORDER BY numerodocumento DESC
            LIMIT 200
        """)
        cols = [desc[0].lower() for desc in cur_aurora.description]
        recientes_rows = cur_aurora.fetchall()
        cur_aurora.close()
        
        a_notificar = []
        for row in recientes_rows:
            record = dict(zip(cols, row))
            doc = str(record.get('numerodocumento') or '').strip()
            if not doc or doc == 'NULL': continue
            
            import re
            def clean_un(val):
                if not val: return ''
                return re.sub(r'(?i)Certificados?\s+de\s+Ventas?\s+Intern[oa]s?\s+para\s+', '', str(val)).strip()

            solicitante = str(record.get('solicitante') or record.get('dim. valor') or '').strip()
            equipo_solicitante = clean_un(solicitante)
            
            sucursal_emisora = str(record.get('unidaddenegocio') or record.get('organizacion') or record.get('empresa') or '').strip()
            sucursal_emisora = clean_un(sucursal_emisora)
            
            # Unidad representativa para el mensaje
            un_principal = equipo_solicitante if equipo_solicitante else sucursal_emisora
                
            desc = str(record.get('documentodescripcion') or record.get('detalledescripcion') or '').strip()
            imp = record.get('importe', 0)
            
            # Buscar el estado actual y el id transaccional
            estado_actual = str(record.get('estadoautorizacion', '')).strip()
            # transaccionid suele ser el campo ID en Finnegans, si no id
            trans_id = record.get('transaccionid') or record.get('id') or ''
            
            # Extraer actividad para botones SOLO si el estado está pendiente (lógica de Compras)
            actividad = ""
            estados_pendientes = ["doc. pendiente firmar", "pendiente de autorización", "autorización", "pendiente"]
            if estado_actual.lower() in estados_pendientes:
                for k in ['actividadworkflow', 'actividad', 'nombreactividad']:
                    if record.get(k) and str(record.get(k)).strip() and str(record.get(k)).strip().upper() not in ('NONE', 'NULL', ''):
                        actividad = str(record.get(k)).strip()
                        break
                
                if not actividad and un_principal:
                    nombre_sucursal = un_principal.replace(' CEE Enriquez', '').replace(' CEE', '').strip()
                    actividad = f"Autoriza N1 por {nombre_sucursal}"
            
            estado_previo = ya_notificados.get(doc)
            es_modificacion = False
            
            if doc not in ya_notificados:
                # Es nuevo
                a_notificar.append((doc, un_principal, equipo_solicitante, sucursal_emisora, desc, imp, estado_actual, trans_id, es_modificacion, actividad))
            elif estado_previo != estado_actual and estado_actual:
                # Modificado (el estado cambió y no está vacío)
                es_modificacion = True
                a_notificar.append((doc, un_principal, equipo_solicitante, sucursal_emisora, desc, imp, estado_actual, trans_id, es_modificacion, actividad))
                
        enviados = 0
        for doc, un_principal, equipo_solicitante, sucursal_emisora, desc, imp_raw, estado, trans_id, es_modificacion, actividad in a_notificar:
            try:
                imp = float(str(imp_raw or '0').replace(',', '.'))
            except:
                imp = 0.0
                
            finnegans_link = ""
            if trans_id:
                # Generar link al comprobante
                finnegans_link = f"https://go.finneg.com/mas/vista?fafViewCode=DF_VIEWER&pk={trans_id}&claseVO=CasoDirectoVO&FAFCLASE_FACADE=FAFTransaccionBSuiteEJB&appitemID=50394"
                
            # Buscar destinatarios (match con Equipo Solicitante o con Sucursal) de forma case-insensitive
            unidades_match = tuple(u.lower() for u in [equipo_solicitante, sucursal_emisora] if u)
            if not unidades_match:
                continue
                
            cur_supa.execute("""
                SELECT DISTINCT u.nombre, u.telegram_chat_id 
                FROM cert_usuarios_unidades cu 
                JOIN cert_usuarios u ON cu.usuario_id = u.id 
                WHERE LOWER(TRIM(cu.unidad_negocio)) IN %s AND cu.notifica_telegram = true AND u.activo = 1
            """, (unidades_match,))
            destinatarios = cur_supa.fetchall()
            
            if destinatarios:
                for nombre, chat_id in destinatarios:
                    if chat_id:
                        telegram_nuevo_certificado(
                            chat_id=chat_id, 
                            comprobante=doc, 
                            descripcion=desc, 
                            unidad=un_principal, 
                            total=imp, 
                            link=APP_URL, 
                            estado=estado, 
                            es_modificacion=es_modificacion, 
                            finnegans_link=finnegans_link,
                            actividad=actividad
                        )
            
            # Marcar o actualizar como notificado
            if not es_modificacion:
                cur_supa.execute("""
                    INSERT INTO cert_notificaciones_log (tipo, comprobante, mensaje, estado, estado_doc)
                    VALUES ('cron_telegram', %s, 'Procesado por cron', 'ok', %s)
                """, (doc, estado))
            else:
                cur_supa.execute("""
                    UPDATE cert_notificaciones_log 
                    SET estado_doc = %s, fecha = CURRENT_TIMESTAMP
                    WHERE comprobante = %s AND tipo = 'cron_telegram'
                """, (estado, doc))
                
            conn_supa.commit()
            enviados += 1
            
        cur_supa.close()
        return {"ok": True, "procesados": enviados, "total_revisados": len(recientes_rows)}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn_supa:
            conn_supa.close()
        if conn_aurora:
            conn_aurora.close()

@app.get("/api/debug-unidades")
def debug_unidades():
    """Endpoint PÚBLICO de diagnóstico para verificar conectividad Aurora y columnas disponibles."""
    import traceback
    result = {"step": "init", "aurora_host": DB_HOST[:40] + "..."}
    try:
        result["step"] = "connecting"
        conn = get_aurora()
        result["step"] = "connected"
        cur = conn.cursor()
        
        # 1. Verificar que la tabla existe y obtener columnas
        cur.execute("SELECT * FROM ceesa_cee_certificados_ventas_internas LIMIT 1")
        columns = [desc[0].lower() for desc in cur.description]
        result["columns_count"] = len(columns)
        result["columns_sample"] = columns[:20]
        
        # 2. Buscar columna de unidad de negocio
        un_col = next((c for c in columns if 'unidad' in c and 'negocio' in c), None)
        result["un_column_found"] = un_col
        
        if not un_col:
            result["error"] = "No se encontró columna con 'unidad' y 'negocio'"
            cur.close(); conn.close()
            return result
        
        # 3. Contar registros totales
        cur.execute(f"SELECT COUNT(*) FROM ceesa_cee_certificados_ventas_internas")
        result["total_rows"] = cur.fetchone()[0]
        
        # 4. Obtener unidades distintas
        query = f"""
            SELECT DISTINCT TRIM(COALESCE({un_col}, '')) as un 
            FROM ceesa_cee_certificados_ventas_internas 
            WHERE {un_col} IS NOT NULL AND TRIM(COALESCE({un_col}, '')) != ''
            ORDER BY un
        """
        cur.execute(query)
        unidades = [row[0] for row in cur.fetchall() if row[0]]
        result["unidades_count"] = len(unidades)
        result["unidades"] = unidades
        result["status"] = "OK"
        
        cur.close()
        conn.close()
        return result
    except Exception as e:
        result["error"] = str(e)
        result["trace"] = traceback.format_exc()[:500]
        return result

@app.get("/api/unidades-negocio")
def list_unidades(user=Depends(get_current_user)):
    """Obtiene las Sucursales y su Empresa Padre desde Aurora."""
    conn = None
    try:
        conn = get_aurora()
        cur = conn.cursor()
        
        # 1. Obtenemos Sucursales de la tabla maestra
        cur.execute("""
            SELECT 
                TRIM(COALESCE(nombreempresa, '')) as sucursal,
                MAX(TRIM(COALESCE(nombreempresapadre, ''))) as empresa_padre
            FROM ceesa_cee_sucursales
            WHERE nombreempresa IS NOT NULL AND TRIM(COALESCE(nombreempresa, '')) != ''
            GROUP BY TRIM(COALESCE(nombreempresa, ''))
        """)
        unidades_lista = [{"sucursal": r[0], "empresa_padre": r[1]} for r in cur.fetchall() if r[0]]
        
        # 2. Obtenemos Equipos Solicitantes (Unidades de Negocio)
        try:
            cur.execute("""
                SELECT DISTINCT TRIM(COALESCE("dim. valor", solicitante, ''))
                FROM ceesa_cee_certificados_ventas_internas
                WHERE COALESCE("dim. valor", solicitante) IS NOT NULL AND TRIM(COALESCE("dim. valor", solicitante)) != ''
            """)
            import re
            vistos = {u["sucursal"].lower() for u in unidades_lista}
            
            for row in cur.fetchall():
                val = row[0]
                val = re.sub(r'(?i)Certificados?\s+de\s+Ventas?\s+Intern[oa]s?\s+para\s+', '', val).strip()
                if val and val.lower() not in vistos:
                    unidades_lista.append({"sucursal": val, "empresa_padre": "Equipos Solicitantes"})
                    vistos.add(val.lower())
        except Exception as e:
            print(f"Error extrayendo Equipos Solicitantes: {e}")
            
        unidades_lista.sort(key=lambda x: x["sucursal"].lower())
        cur.close()
        return {"data": unidades_lista}
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Error en list_unidades: {e}\n{tb}")
        return {"data": [], "error": str(e), "trace": tb[:300]}
    finally:
        if conn:
            conn.close()

# ═══════════════════════════════════════════════════════
# NOTIFICACIONES TELEGRAM
# ═══════════════════════════════════════════════════════

@app.post("/api/notificar-certificado/{comprobante}")
def notificar_certificado(comprobante: str, user=Depends(get_current_user)):
    """Envía notificación Telegram a los responsables de la UN del certificado."""
    # 1. Obtener datos del certificado desde Aurora
    conn_aurora = None
    conn_supa = None
    try:
        conn_aurora = get_aurora()
        cur = conn_aurora.cursor()
        cur.execute("""
            SELECT *
            FROM ceesa_cee_certificados_ventas_internas 
            WHERE numerodocumento = %s
            ORDER BY CAST(NULLIF(importe,'0') AS DECIMAL) DESC NULLS LAST
            LIMIT 1
        """, (comprobante,))
        columns = [desc[0].lower() for desc in cur.description]
        row = cur.fetchone()
        cur.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="Certificado no encontrado")
            
        record = dict(zip(columns, row))
        import re
        def clean_un(val):
            if not val: return ''
            return re.sub(r'(?i)Certificados?\s+de\s+Ventas?\s+Intern[oa]s?\s+para\s+', '', str(val)).strip()

        solicitante = str(record.get('solicitante') or record.get('dim. valor') or '').strip()
        equipo_solicitante = clean_un(solicitante)
        
        sucursal_emisora = str(record.get('unidaddenegocio') or record.get('organizacion') or record.get('empresa') or '').strip()
        sucursal_emisora = clean_un(sucursal_emisora)
        
        un_principal = equipo_solicitante if equipo_solicitante else sucursal_emisora
        
        descripcion = str(record.get('documentodescripcion') or record.get('detalledescripcion') or '').strip()
        try:
            total = float(record.get('importe') or 0)
        except:
            total = 0.0
        
        unidades_match = tuple(u.lower() for u in [equipo_solicitante, sucursal_emisora] if u)
        if not unidades_match:
            return {"ok": False, "error": "No se determinó unidad destino ni sucursal"}
            
        # 2. Buscar responsables de esa UN con Telegram activo de forma case-insensitive
        conn_supa = get_supabase()
        cur2 = conn_supa.cursor()
        cur2.execute("""
            SELECT DISTINCT u.nombre, u.telegram_chat_id 
            FROM cert_usuarios u
            JOIN cert_usuarios_unidades uu ON u.id = uu.usuario_id
            WHERE LOWER(TRIM(uu.unidad_negocio)) IN %s 
              AND uu.notifica_telegram = true 
              AND u.telegram_chat_id IS NOT NULL 
              AND u.telegram_chat_id != ''
              AND u.activo = 1
        """, (unidades_match,))
        destinatarios = cur2.fetchall()
        
        resultados = []
        for nombre, chat_id in destinatarios:
            res = telegram_nuevo_certificado(
                chat_id=chat_id,
                comprobante=comprobante,
                descripcion=descripcion,
                unidad=un_principal,
                total=total,
                link=APP_URL
            )
            # Log de notificación
            cur2.execute("""
                INSERT INTO cert_notificaciones_log (tipo, destinatario, comprobante, mensaje, estado)
                VALUES ('telegram', %s, %s, %s, %s)
            """, (nombre, comprobante, f"Nuevo certificado {comprobante}", "ok" if res.get("ok") else "error"))
            resultados.append({"nombre": nombre, "ok": res.get("ok", False)})
        
        conn_supa.commit()
        cur2.close()
        
        return {"ok": True, "notificados": len(resultados), "detalle": resultados}
    finally:
        if conn_aurora:
            conn_aurora.close()
        if conn_supa:
            conn_supa.close()

@app.post("/api/telegram-test/{user_id}")
def test_telegram(user_id: int, admin=Depends(require_admin)):
    """Envía un mensaje de prueba al Telegram del usuario."""
    conn = get_supabase()
    try:
        cur = conn.cursor()
        cur.execute("SELECT telegram_chat_id FROM cert_usuarios WHERE id = %s", (user_id,))
        row = cur.fetchone()
        cur.close()
        
        if not row or not row[0]:
            raise HTTPException(status_code=400, detail="Usuario sin Chat ID de Telegram")
        
        result = telegram_test(row[0])
        return result
    finally:
        conn.close()

# ═══════════════════════════════════════════════════════
# DASHBOARD — INDICADORES (Mantener endpoint existente)
# ═══════════════════════════════════════════════════════

@app.get("/api/debug")
def debug_endpoint():
    conn = get_aurora()
    cur = conn.cursor()
    cur.execute("""
        SELECT numerodocumento, importe, producto, detalledescripcion, precio, cantidadworkflow
        FROM ceesa_cee_certificados_ventas_internas 
        WHERE numerodocumento = 'CI-0001-00000003'
        ORDER BY importe DESC NULLS LAST
    """)
    cols = [desc[0] for desc in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"sample_rows": rows}

@app.get("/api/indicadores")
def get_indicadores(user=Depends(get_current_user)):
    conn = None
    try:
        # 1. Obtener sucursales permitidas si no es admin
        unidades_permitidas = set()
        if user.get("rol") != "admin":
            conn_supa = get_supabase()
            cur_supa = conn_supa.cursor()
            cur_supa.execute("SELECT unidad_negocio FROM cert_usuarios_unidades WHERE usuario_id = %s", (user.get('id') or user.get('sub'),))
            unidades_permitidas = {r[0].strip() for r in cur_supa.fetchall() if r[0]}
            cur_supa.close()
            conn_supa.close()

        conn = get_aurora()
        
        query = "SELECT * FROM ceesa_cee_certificados_ventas_internas ORDER BY 1 DESC LIMIT 50000"
        
        cursor = conn.cursor()
        cursor.execute(query)
        columns_db = [desc[0] for desc in cursor.description]
        data_rows = cursor.fetchall()
        cursor.close()
        
        # ─── PASO 1: Agrupar por comprobante ───
        comprobantes = {}
        for i, row in enumerate(data_rows):
            record = dict(zip(columns_db, row))
            
            import re
            def normalize_match(s):
                if not s: return ""
                return re.sub(r'[^a-zA-Z0-9]', '', str(s)).upper()

            # Filtro estricto por Sucursal/Prestador
            # Usamos 'empresa' porque allí Finnegans guarda la Sucursal (ej: Informática y Tecnología CEE ENRIQUEZ)
            empresa_val = str(record.get('empresa', '') or '').strip()
            
            if user.get("rol") != "admin":
                match_found = False
                norm_empresa = normalize_match(empresa_val)
                for u in unidades_permitidas:
                    if norm_empresa == normalize_match(u):
                        match_found = True
                        break
                if not match_found:
                    continue

            num_doc = record.get('comprobante', '')
            if not num_doc or num_doc == 'NULL':
                num_doc = record.get('documento', '')
                if not num_doc or num_doc == 'NULL':
                    num_doc = f"ROW-{i}"
                
            imp_raw = record.get('total', '0') or '0'
            try:
                imp = float(str(imp_raw).replace(',', '.'))
            except:
                imp = 0.0
                
            gravado_raw = record.get('gravado', '0') or '0'
            try:
                gravado = float(str(gravado_raw).replace(',', '.'))
            except:
                gravado = 0.0
            
            producto = str(record.get('productonombre', '') or '').strip()
            
            if num_doc not in comprobantes:
                comprobantes[num_doc] = {
                    'metadata': record,
                    'max_importe': imp,
                    'max_gravado': gravado,
                    'items': [],
                }
            
            if imp > comprobantes[num_doc]['max_importe']:
                comprobantes[num_doc]['max_importe'] = imp
            if gravado > comprobantes[num_doc]['max_gravado']:
                comprobantes[num_doc]['max_gravado'] = gravado
            
            if producto and producto != 'NULL':
                cant_raw = record.get('itemcantidad', '0') or '0'
                precio_raw = record.get('itemprecio', '0') or '0'
                itemimp_raw = record.get('itemimporte', '0') or '0'
                try:
                    cant = float(str(cant_raw).replace(',', '.'))
                except:
                    cant = 1.0
                try:
                    precio = float(str(precio_raw).replace(',', '.'))
                except:
                    precio = 0.0
                try:
                    itemimp = float(str(itemimp_raw).replace(',', '.'))
                except:
                    itemimp = 0.0
                
                item_data = {
                    'Producto': producto,
                    'Cantidad': cant,
                    'Precio': precio,
                    'Importe': itemimp,
                    'Unidad': record.get('unidadnombre', '')
                }
                if item_data not in comprobantes[num_doc]['items']:
                    comprobantes[num_doc]['items'].append(item_data)
        
        # ─── PASO 2: Construir registros finales ───
        records = []
        for num_doc, data in comprobantes.items():
            meta = data['metadata']
            
            fecha_val = meta.get('fecha', '')
            if not fecha_val or fecha_val == 'NULL':
                fecha_val = meta.get('fechaalta', '')
            fecha_raw = fecha_val
            fecha_fmt = ''
            if fecha_raw:
                try:
                    if isinstance(fecha_raw, datetime):
                        fecha_fmt = fecha_raw.strftime('%d/%m/%Y')
                    elif isinstance(fecha_raw, str) and len(fecha_raw) >= 10:
                        fecha_fmt = datetime.strptime(fecha_raw[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
                    else:
                        fecha_fmt = str(fecha_raw)
                except:
                    fecha_fmt = str(fecha_raw)
            
            def clean(val):
                if val is None or val == 'NULL':
                    return ''
                return str(val).strip()
            
            desc = clean(meta.get('descripcion', ''))
            items_list = data['items']
            
            total_val = data['max_importe']
            gravado_val = data['max_gravado']
            iva_val = round(total_val - gravado_val, 2)
            
            record = {
                'Fecha': fecha_fmt,
                'Comprobante': num_doc,
                'Empresa': clean(meta.get('empresa', '')),
                'Cliente': clean(meta.get('cliente', '')),
                'Descripción': desc,
                'Solicitante': clean(meta.get('equiposolicitantenombre', '')),
                'EstadoAutorizacion': 'Autorizado',
                'Total Bruto': str(total_val),
                'Neto Gravado': str(gravado_val),
                'IVA': str(iva_val),
                'UnidadNegocio': clean(meta.get('unidadnombre', '')),
                'Concepto': clean(meta.get('productonombre', '')),
                'items': items_list,
            }
            records.append(record)
        

        # --- PASO 2.5: Añadir Ajustes Excel (Ingresos) ---
        try:
            conn_supa = get_supabase()
            cur_supa = conn_supa.cursor()
            
            # Construir la query base
            query_ajustes = "SELECT id, tipo_movimiento, categoria, fecha_carga, concepto, observaciones, importe, unidad_negocio, periodo FROM cert_ajustes_excel WHERE tipo_movimiento = 'INGRESO'"
            params_ajustes = []
            
            if user.get("rol") != "admin" and unidades_permitidas:
                # Filtrar por unidades permitidas
                placeholders = ','.join(['%s'] * len(unidades_permitidas))
                query_ajustes += f" AND unidad_negocio IN ({placeholders})"
                params_ajustes.extend(list(unidades_permitidas))
            
            cur_supa.execute(query_ajustes, tuple(params_ajustes))
            rows_ajustes = cur_supa.fetchall()
            
            for r in rows_ajustes:
                id_ajuste = r[0]
                categoria = r[2] or "Ajuste Manual"
                periodo_val = str(r[8] or "")
                
                # Use the first day of the periodo so the frontend extracts the correct period
                if periodo_val and "/" in periodo_val:
                    fecha = f"01/{periodo_val}"
                else:
                    fecha = str(r[3]) if r[3] else ""
                    if fecha and len(fecha) >= 10:
                        try:
                            # Ya tenemos from datetime import datetime importado a nivel global
                            fecha = datetime.strptime(fecha[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
                        except:
                            pass
                            
                concepto_str = f"[{r[7]}] {r[4]}"
                observaciones = r[5] or "-"
                importe_val = float(r[6] or 0)
                
                record_ajuste = {
                    'Fecha': fecha,
                    'Comprobante': r[5] if r[5] else f"EXCEL-{id_ajuste}",
                    'id_ajuste': id_ajuste,
                    'Empresa': r[7],  # unidad_negocio
                    'Cliente': '-',
                    'Descripción': r[4], # concepto
                    'Solicitante': r[7], # unidad_negocio
                    'EstadoAutorizacion': 'Ajuste',
                    'Total Bruto': str(importe_val),
                    'Neto Gravado': str(importe_val),
                    'IVA': '0.0',
                    'UnidadNegocio': r[7],
                    'Concepto': r[4], # concepto
                    'origen': 'AJUSTE EXCEL',
                    'items': [{
                        'id': f'item-{id_ajuste}',
                        'producto': r[4],
                        'descripcion': r[4],
                        'precio': importe_val,
                        'cantidad': 1,
                        'total': importe_val,
                        'Importe': importe_val
                    }]
                }
                records.append(record_ajuste)
            
            cur_supa.close()
            conn_supa.close()
        except Exception as e:
            print(f"Error consultando Ajustes Excel en indicadores: {e}")

        if not records:

            raise Exception("No data found en Aurora")
            
        final_columns = ['Fecha', 'Comprobante', 'Empresa', 'Cliente', 'Descripción', 'Solicitante', 'EstadoAutorizacion', 'Neto Gravado', 'IVA', 'Total Bruto', 'UnidadNegocio', 'Concepto']
        
    except Exception as e:
        print(f"Error consultando BD: {e}")
        records = [
            {'Fecha': '01/05/2026', 'Comprobante': 'CVI-0001', 'Descripción': 'Mock', 'Empresa': 'CEE', 'Total Bruto': '0', 'Neto Gravado': '0', 'IVA': '0', 'EstadoAutorizacion': 'Sin Estado', 'Solicitante': '', 'UnidadNegocio': '', 'items': []}
        ]
        final_columns = list(records[0].keys())

    finally:
        if conn:
            conn.close()

    # ─── PASO 3: KPIs ───
    sum_neto_gravado = 0.0
    sum_iva = 0.0
    sum_total = 0.0
    clientes_set = set()
    
    for record in records:
        try:
            sum_total += float(record.get('Total Bruto', '0') or '0')
            sum_neto_gravado += float(record.get('Neto Gravado', '0') or '0')
            sum_iva += float(record.get('IVA', '0') or '0')
        except:
            pass
        empresa = record.get('Empresa', '')
        if empresa:
            clientes_set.add(empresa)
    
    return {
        "kpis": {
            "total_certificados": len(records),
            "neto_gravado": sum_neto_gravado,
            "iva": sum_iva,
            "total_final": sum_total,
            "clientes_activos": len(clientes_set)
        },
        "columns": final_columns,
        "data": records
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


class PresentarPeriodoReq(BaseModel):
    unidad_negocio: str
    periodo: str

@app.post("/api/cierre/presentar")
def presentar_periodo(req: PresentarPeriodoReq, current_user = Depends(get_current_user)):
    """Toma la info en vivo y la guarda en las tablas de snapshot."""
    try:
        req.periodo = normalize_periodo(req.periodo)
        # Calcular en vivo usando la misma logica (reutilizamos la que usaremos para get_informe_mensual)
        informe = get_informe_mensual_calculo_vivo(req.unidad_negocio, req.periodo)
        
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        # Verificar si ya existe
        cur_supa.execute("SELECT id FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s", (req.unidad_negocio, req.periodo))
        if cur_supa.fetchone():
            cur_supa.close()
            conn_supa.close()
            raise HTTPException(status_code=400, detail="El periodo ya se encuentra presentado.")
            
        cur_supa.execute(
            "INSERT INTO cert_cierres_mensuales (unidad_negocio, periodo, usuario_cierre, fecha_cierre) VALUES (%s, %s, %s, NOW()) RETURNING id",
            (req.unidad_negocio, req.periodo, current_user.get("sub", current_user.get("email", "admin")))
        )
        cierre_id = cur_supa.fetchone()[0]
        
        for item in informe["ingresos"] + informe["gastos"]:
            cur_supa.execute(
                "INSERT INTO cert_cierres_detalle (cierre_id, origen, tipo_movimiento, categoria, fecha, concepto, comprobante, proveedor, importe) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (cierre_id, item["origen"], item["tipo_movimiento"], item["categoria"], item["fecha"] if item["fecha"] else None, item["concepto"], item["comprobante"], item.get("proveedor"), item["importe"])
            )
            
        conn_supa.commit()
        cur_supa.close()
        conn_supa.close()
        return {"status": "ok", "cierre_id": cierre_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ReabrirPeriodoReq(BaseModel):
    unidad_negocio: str
    periodo: str

@app.post("/api/cierre/reabrir")
def reabrir_periodo(req: ReabrirPeriodoReq, current_user = Depends(get_current_user)):
    """Solo permite reabrir al administrador, borrando el snapshot."""
    if current_user.get("role", "admin") != "admin":
         raise HTTPException(status_code=403, detail="Solo un administrador puede reabrir periodos presentados.")
    try:
        req.periodo = normalize_periodo(req.periodo)
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        cur_supa.execute("SELECT id FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s", (req.unidad_negocio, req.periodo))
        row = cur_supa.fetchone()
        if not row:
            cur_supa.close()
            conn_supa.close()
            raise HTTPException(status_code=400, detail="El periodo no se encuentra presentado.")
            
        cierre_id = row[0]
        cur_supa.execute("DELETE FROM cert_cierres_detalle WHERE cierre_id = %s", (cierre_id,))
        cur_supa.execute("DELETE FROM cert_cierres_mensuales WHERE id = %s", (cierre_id,))
        
        conn_supa.commit()
        cur_supa.close()
        conn_supa.close()
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def get_informe_mensual_calculo_vivo(unidad_negocio: str, periodo: str):
    periodo = normalize_periodo(periodo)
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    
    # 1. Config Unidades
    # La unidad_negocio que llega ya es la sucursal de Finnegans
    sucursales = [unidad_negocio]
    cur_supa.execute("SELECT centro_id FROM cert_config_centros_costo WHERE sucursal = %s", (unidad_negocio,))
    centros = [r[0] for r in cur_supa.fetchall()]

    # 1b. Configuración Avanzada para esta Sucursal
    cur_supa.execute("SELECT codigo FROM cert_config_ingresos_comprobantes WHERE sucursal = %s", (unidad_negocio,))
    ingresos_subtipos = [r[0] for r in cur_supa.fetchall()]

    cur_supa.execute("SELECT tipo_asiento_id FROM cert_config_gastos_asientos WHERE sucursal = %s", (unidad_negocio,))
    gastos_asientos_ids = [r[0] for r in cur_supa.fetchall()]

    cur_supa.execute("SELECT subtipo_id FROM cert_config_gastos_compras WHERE sucursal = %s", (unidad_negocio,))
    gastos_compras_ids = [r[0] for r in cur_supa.fetchall()]

    # 2. Config Categorías Gastos
    cur_supa.execute("SELECT categoria, cuenta_codigo FROM cert_config_gastos_cuentas")
    cat_rows = cur_supa.fetchall()
    cat_map = {r[1]: r[0] for r in cat_rows}

    # 3. Excel Items
    y, m = periodo.split('-')
    cur_supa.execute(
        "SELECT tipo, fecha, concepto, importe FROM cert_excel_items WHERE unidad_negocio = %s AND EXTRACT(YEAR FROM fecha) = %s AND EXTRACT(MONTH FROM fecha) = %s",
        (unidad_negocio, int(y), int(m))
    )
    excel_items = cur_supa.fetchall()
    
    cur_supa.close()
    conn_supa.close()

    ingresos = []
    gastos = []
    
    for ex in excel_items:
        item = {
            "origen": "EXCEL",
            "tipo_movimiento": ex[0],
            "categoria": "Ingresos adicionales" if ex[0].upper() == 'INGRESO' else "Egresos adicionales",
            "fecha": str(ex[1]),
            "concepto": ex[2],
            "comprobante": "EXCEL",
            "importe": float(ex[3])
        }
        if ex[0].upper() == 'INGRESO':
            ingresos.append(item)
        else:
            gastos.append(item)

    # 4. Aurora Finnegans
    conn = get_aurora()
    cur = conn.cursor()
    
    # -- INGRESOS --
    # Nota: Corregimos los nombres de columnas para evitar el error "column producto does not exist"
    # Usamos documento y comprobante, productonombre, y equiposolicitantenombre
    sql_ingresos = """
    SELECT fecha, documento, comprobante, productonombre, equiposolicitantenombre, itemimporte, itemimportegravado, itemcantidad
    FROM ceesa_cee_certificados_ventas_internas
    WHERE EXTRACT(YEAR FROM CAST(fecha AS TIMESTAMP)) = %s 
      AND EXTRACT(MONTH FROM CAST(fecha AS TIMESTAMP)) = %s
      AND productonombre IS NOT NULL AND TRIM(productonombre) != ''
    """
    params_ingresos = [int(y), int(m)]
    
    # Aplicar filtro de unidad
    if sucursales:
        sucs_upper = ",".join(f"UPPER('{s}')" for s in sucursales)
        sql_ingresos += f" AND UPPER(empresa) IN ({sucs_upper})" # 'empresa' es el prestador
        
    # Aplicar filtro de subtipos de comprobante de ingresos si hay configurados para esta sucursal
    if ingresos_subtipos:
        subtipos_str = ",".join(f"'{s}'" for s in ingresos_subtipos)
        sql_ingresos += f" AND (split_part(comprobante, '-', 1) IN ({subtipos_str}) OR TRIM(split_part(documento, '-', 1)) IN ({subtipos_str}))"
    
    try:
        cur.execute(sql_ingresos, params_ingresos)
        rows_ingresos = cur.fetchall()
        for r in rows_ingresos:
            val_gravado = safe_float(r[6]) if len(r) > 6 else 0.0
            val_total = safe_float(r[5]) if len(r) > 5 else 0.0
            importe_ingreso = val_gravado if val_gravado != 0 else val_total
            
            comp_val = r[2] if r[2] and str(r[2]).strip() else r[1] # fallback a documento si comprobante es vacio
            cantidad = safe_float(r[7]) if len(r) > 7 else 0.0
            
            ingresos.append({
                "origen": "FINNEGANS",
                "tipo_movimiento": "INGRESO",
                "categoria": "Ventas Internas",
                "fecha": str(r[0]) if r[0] else None,
                "concepto": r[3] or "Sin Detalle",        # productonombre
                "comprobante": comp_val or "N/A",         # comprobante o documento
                "proveedor": r[4] or "-",                 # equiposolicitantenombre (UI usa proveedor/cliente)
                "cantidad": cantidad,
                "importe": importe_ingreso
            })
    except Exception as e:
        print("Error Ingresos:", e)
        ingresos.append({"origen": "ERROR", "tipo_movimiento": "INGRESO", "categoria": "ERROR", "fecha": None, "concepto": "ERROR: " + str(e), "comprobante": "N/A", "proveedor": "-", "cantidad": 0, "importe": 0})

    # Obtener empresa padre
    conn_a_tmp = get_aurora()
    cur_a_tmp = conn_a_tmp.cursor()
    cur_a_tmp.execute("""
        SELECT MAX(TRIM(COALESCE(nombreempresapadre, '')))
        FROM ceesa_cee_sucursales
        WHERE TRIM(COALESCE(nombreempresa, '')) = %s
    """, (unidad_negocio,))
    padre_row = cur_a_tmp.fetchone()
    empresa_padre = padre_row[0] if padre_row and padre_row[0] else unidad_negocio
    cur_a_tmp.close()
    conn_a_tmp.close()

    empresas_validas = list(set([unidad_negocio, empresa_padre]))
    empresas_str = ",".join(f"'{e}'" for e in empresas_validas)
    
    where_gastos = [f"FAFEmpresa.nombre IN ({empresas_str})"]
    if centros:
        ccs_str = ",".join(f"'{c}'" for c in centros)
        where_gastos.append(f"BSTransaccionDimension.registroid IN ({ccs_str})")
        
    cond_gastos = f"({ ' AND '.join(where_gastos) })" if where_gastos else "1=1"
    
    # Filtros para Asientos
    where_subtipos_asientos = []
    if gastos_asientos_ids:
        asientos_str = ",".join(f"'{a}'" for a in gastos_asientos_ids)
        where_subtipos_asientos.append(f"FAFTransaccionCategoria.transaccioncategoriaid IN ({asientos_str})")
    
    # === 1. Asientos Contables ===
    if where_subtipos_asientos:
        cond_asientos = f"({ ' OR '.join(where_subtipos_asientos) })"
        sql_asientos = f"""
        SELECT
            BSCuenta.codigo AS CuentaCodigo,
            BSCuenta.nombre AS Cuenta,
            CAST(BSAsientoItem.fecha AS TIMESTAMP) AS Fecha,
            FAFEmpresa.nombre AS Sucursal,
            COALESCE(CAST(BSTransaccionDimension.importemonprincipal AS NUMERIC), CAST(BSAsientoItem.importemonprincipal AS NUMERIC)) * CAST(BSAsientoItem.debehaber AS NUMERIC) AS Importe,
            BSTransaccion.descripcion AS Comprobante,
            VProv.nombre AS Proveedor
        FROM ceesa_bsasientoitem AS BSAsientoItem
        INNER JOIN ceesa_bscuenta AS BSCuenta ON BSAsientoItem.cuentaid = BSCuenta.cuentaid
        INNER JOIN ceesa_bstransaccion AS BSTransaccion ON BSAsientoItem.transaccionid = BSTransaccion.transaccionid
        INNER JOIN ceesa_fafempresa AS FAFEmpresa ON BSTransaccion.empresaid = FAFEmpresa.empresaid
        LEFT JOIN ceesa_bstransacciondimension AS BSTransaccionDimension 
            ON BSAsientoItem.asientoitemid = BSTransaccionDimension.asientoitemid 
            AND BSTransaccion.transaccionid = BSTransaccionDimension.transaccionid
            AND BSTransaccionDimension.dimensionid = '999999'
        LEFT JOIN ceesa_bscentrocosto AS BSCentroCosto ON BSTransaccionDimension.registroid = BSCentroCosto.centrocostoid
        INNER JOIN ceesa_faftransaccionsubtipo AS FAFTransaccionSubtipo ON BSTransaccion.transaccionsubtipoid = FAFTransaccionSubtipo.transaccionsubtipoid
        INNER JOIN ceesa_faftransaccioncategoria AS FAFTransaccionCategoria ON FAFTransaccionSubtipo.transaccioncategoriaid = FAFTransaccionCategoria.transaccioncategoriaid
        LEFT JOIN ceesa_vprov AS VProv ON BSTransaccion.entidadid = VProv.entidadid
        WHERE FAFEmpresa.empresaidpadre = '1'
          AND BSCuenta.impactaresultados = '1'
          AND EXTRACT(YEAR FROM CAST(BSAsientoItem.fecha AS TIMESTAMP)) = %s
          AND EXTRACT(MONTH FROM CAST(BSAsientoItem.fecha AS TIMESTAMP)) = %s
          AND ({cond_gastos})
          AND ({cond_asientos})
        """
        try:
            cur.execute(sql_asientos, [int(y), int(m)])
            rows_asientos = cur.fetchall()
            for r in rows_asientos:
                cat = cat_map.get(r[0], r[1])
                gastos.append({
                    "origen": "FINNEGANS",
                    "tipo_movimiento": "EGRESO",
                    "categoria": cat,
                    "fecha": str(r[2]) if r[2] else None,
                    "concepto": f"{r[0]} - {r[1]}",
                    "comprobante": r[5] or "Asiento Contable",
                    "proveedor": r[6],
                    "importe": float(r[4] or 0)
                })
        except Exception as e:
            print("Error Gastos Asientos:", e)

    # === 2. Comprobantes de Compra (Nuevo Dataset) ===
    if gastos_compras_ids:
        try:
            compras_id_str = ",".join(f"'{c}'" for c in gastos_compras_ids)
            cur.execute(f"SELECT nombre FROM ceesa_faftransaccionsubtipo WHERE transaccionsubtipoid IN ({compras_id_str})")
            compras_nombres = [r[0] for r in cur.fetchall()]
            
            if compras_nombres:
                compras_nombres_str = ",".join(f"'{n}'" for n in compras_nombres)
                
                where_compras = []
                if centros:
                    ccs_str = ",".join(f"'{c}'" for c in centros)
                    where_compras.append(f"centrocostoid IN ({ccs_str})")
                cond_compras = f"({ ' AND '.join(where_compras) })" if where_compras else "1=1"
                
                sql_compras = f"""
                SELECT
                    CAST(fecha AS TIMESTAMP) AS Fecha,
                    tipodocumento AS Categoria,
                    proveedor AS Concepto,
                    numerodocumento AS Comprobante,
                    SUM(CAST(importeimputado AS NUMERIC)) AS Importe,
                    centrocosto AS Sucursal
                FROM ceesa_cee_gastos_cc
                WHERE EXTRACT(YEAR FROM CAST(fecha AS TIMESTAMP)) = %s
                  AND EXTRACT(MONTH FROM CAST(fecha AS TIMESTAMP)) = %s
                  AND ({cond_compras})
                  AND tipodocumento IN ({compras_nombres_str})
                GROUP BY
                    fecha,
                    tipodocumento,
                    proveedor,
                    numerodocumento,
                    centrocosto
                """
                cur.execute(sql_compras, [int(y), int(m)])
                rows_compras = cur.fetchall()
                for r in rows_compras:
                    gastos.append({
                        "origen": "FINNEGANS",
                        "tipo_movimiento": "EGRESO",
                        "categoria": "Costos",  # O r[1] si queremos que sea el tipo de documento
                        "fecha": str(r[0]) if r[0] else None,
                        "concepto": f"{r[1]} - {r[2] or 'Sin Proveedor'}",
                        "comprobante": r[3] or "S/N",
                        "proveedor": r[2],
                        "importe": float(r[4] or 0)
                    })
        except Exception as e:
            print("Error Gastos Compras:", e)

        # === 3. Consumos de Inventario (Consolidado) ===
        try:
            cur.execute("""
                SELECT 
                    COUNT(*),
                    SUM(CASE WHEN importevalorizadoconsumoprod IS NOT NULL AND importevalorizadoconsumoprod != 'NULL' AND importevalorizadoconsumoprod != '' 
                             THEN CAST(importevalorizadoconsumoprod AS DOUBLE PRECISION) 
                             ELSE 0.0 END)
                FROM analisis_de_consumos_de_produccion
                WHERE (empresa = %s OR depositoorigenconsumoprod ILIKE %s OR depositoorigenconsumoprod = %s)
                  AND SUBSTRING(fecha, 1, 7) = %s
            """, (unidad_negocio, f"%{unidad_negocio}%", unidad_negocio, periodo))
            row_c = cur.fetchone()
            if row_c and (row_c[0] or 0) > 0:
                total_consumos = float(row_c[1] or 0.0)
                y_part, m_part = periodo.split('-')
                comprobante_label = f"Consumo de Insumos ({m_part}/{y_part})"
                gastos.append({
                    "origen": "CONSUMOS",
                    "tipo_movimiento": "EGRESO",
                    "categoria": "Materiales",
                    "fecha": f"{periodo}-01",
                    "concepto": "Consumo de Insumos de Producción / Inventario (Soporte)",
                    "comprobante": comprobante_label,
                    "proveedor": "Consumos de Producción",
                    "importe": total_consumos
                })
        except Exception as e:
            print("Error calculando total consumos consolidado:", e)

    cur.close()
    conn.close()
    
    try:
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        # Convertir periodo YYYY-MM a MM/YYYY para consultar cert_ajustes_excel
        periodo_mmyyyy = periodo
        if "-" in periodo:
            y_part, m_part = periodo.split('-')
            periodo_mmyyyy = f"{m_part}/{y_part}"
            
        if unidad_negocio == "Todas":
            cur_supa.execute("SELECT id, tipo_movimiento, categoria, fecha_carga, concepto, observaciones, importe, unidad_negocio FROM cert_ajustes_excel WHERE periodo = %s", (periodo_mmyyyy,))
        else:
            cur_supa.execute("SELECT id, tipo_movimiento, categoria, fecha_carga, concepto, observaciones, importe, unidad_negocio FROM cert_ajustes_excel WHERE unidad_negocio = %s AND periodo = %s", (unidad_negocio, periodo_mmyyyy))
            
        rows_ajustes = cur_supa.fetchall()
        for r in rows_ajustes:
            db_cat = r[2] or "Ajuste Manual"
            if db_cat == "Ajuste Excel":
                categoria_label = "Ingresos adicionales" if r[1] == "INGRESO" else "Egresos adicionales"
            elif db_cat == "Gastos de Compra":
                categoria_label = "Costos"
            else:
                categoria_label = db_cat
            item = {
                "origen": "AJUSTE EXCEL",
                "id_ajuste": r[0],
                "tipo_movimiento": "INGRESO" if r[1] == "INGRESO" else "EGRESO",
                "categoria": categoria_label,
                "fecha": str(r[3]) if r[3] else None,
                "concepto": f"[{r[7]}] {r[4]}",
                "comprobante": r[5] or "-",
                "importe": float(r[6] or 0)
            }
            if r[1] == "INGRESO":
                ingresos.append(item)
            else:
                gastos.append(item)
        cur_supa.close()
        conn_supa.close()
    except Exception as e:
        print("Error Ajustes Excel:", e)
        
    # === 4. Incorporar Sueldos de RRHH ===
    try:
        rrhh_data = get_rrhh(empresa=unidad_negocio, periodo=periodo)
        costo_empresa = rrhh_data.get("totales", {}).get("costo_empresa", 0.0)
        if costo_empresa > 0:
            gastos.append({
                "origen": "RRHH",
                "tipo_movimiento": "EGRESO",
                "categoria": "Sueldos y Cargas Sociales",
                "fecha": f"{periodo}-28",
                "concepto": "Liquidación de Sueldos - RRHH",
                "comprobante": f"RRHH-{periodo}",
                "proveedor": "Recursos Humanos",
                "importe": float(costo_empresa)
            })
    except Exception as e:
        print("Error al incorporar sueldos de RRHH en calculo vivo:", e)
        
    # === 5. Incorporar Certificados de Obras (en Ingresos) ===
    try:
        obras = get_certificados_obras_live(unidad_negocio, periodo)
        for ob in obras:
            if ob.get("estado") == "CONFIRMADO":
                # Sumar el total de la columna parcial_presente de todos los ítems del certificado
                total_parcial_presente = sum(float(item.get("parcial_presente") or 0.0) for item in ob.get("items", []))
                if total_parcial_presente > 0:
                    ingresos.append({
                        "origen": "OBRAS",
                        "tipo_movimiento": "INGRESO",
                        "categoria": "Certificados de Obras",
                        "fecha": ob.get("fecha_certificado")[:10] if ob.get("fecha_certificado") else f"{periodo}-01",
                        "concepto": f"Certificado de Obra: {ob.get('obra') or 'Sin nombre'}",
                        "comprobante": f"Certificado #{ob.get('numero_interno')}",
                        "proveedor": ob.get("comitente") or "-",
                        "importe": total_parcial_presente
                    })
    except Exception as e:
        print("Error al incorporar certificados de obras en calculo vivo:", e)
    
    return {
        "ingresos": ingresos,
        "gastos": gastos
    }

def get_informe_mensual_ingresos(unidad_negocio: str, periodo: str):
    data = get_informe_mensual_calculo_vivo(unidad_negocio, periodo)
    return data["ingresos"]

def get_informe_totales(unidad_negocio: str, periodo: str):
    data = get_informe_mensual_calculo_vivo(unidad_negocio, periodo)
    ingresos = data["ingresos"]
    gastos = data["gastos"]
    return {
        "ingresos": sum(i["importe"] for i in ingresos),
        "gastos": sum(g["importe"] for g in gastos),
        "neto": sum(i["importe"] for i in ingresos) - sum(g["importe"] for g in gastos)
    }

@app.get("/api/informes/mensual")
def get_informe_mensual(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    try:
        periodo = normalize_periodo(periodo)
        conn_supa = get_supabase()
        cur_supa = conn_supa.cursor()
        
        cur_supa.execute(
            "SELECT id, usuario_cierre, fecha_cierre FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s",
            (unidad_negocio, periodo)
        )
        cierre = cur_supa.fetchone()
        
        if cierre:
            cierre_id = cierre[0]
            cur_supa.execute(
                "SELECT origen, tipo_movimiento, categoria, fecha, concepto, comprobante, proveedor, importe FROM cert_cierres_detalle WHERE cierre_id = %s",
                (cierre_id,)
            )
            rows = cur_supa.fetchall()
            
            ingresos = []
            gastos = []
            for r in rows:
                db_cat = r[2] or "Ajuste Manual"
                if db_cat == "Ajuste Excel":
                    categoria_label = "Ingresos adicionales" if r[1] == "INGRESO" else "Egresos adicionales"
                elif db_cat == "Gastos de Compra":
                    categoria_label = "Costos"
                else:
                    categoria_label = db_cat
                item = {
                    "origen": r[0],
                    "tipo_movimiento": r[1],
                    "categoria": categoria_label,
                    "fecha": str(r[3]) if r[3] else None,
                    "concepto": r[4],
                    "comprobante": r[5],
                    "proveedor": r[6],
                    "importe": float(r[7])
                }
                if r[1] == "INGRESO":
                    ingresos.append(item)
                else:
                    gastos.append(item)
                    
            cur_supa.close()
            conn_supa.close()
            
            return {
                "estado_cierre": "CERRADO",
                "usuario_cierre": cierre[1],
                "fecha_cierre": str(cierre[2]),
                "ingresos": ingresos,
                "gastos": gastos,
                "totales": {
                    "ingresos": sum(i["importe"] for i in ingresos),
                    "gastos": sum(g["importe"] for g in gastos)
                }
            }
            
        cur_supa.close()
        conn_supa.close()
        
        # En vivo
        data = get_informe_mensual_calculo_vivo(unidad_negocio, periodo)
        return {
            "estado_cierre": "ABIERTO",
            "usuario_cierre": None,
            "fecha_cierre": None,
            "ingresos": data["ingresos"],
            "gastos": data["gastos"],
            "totales": {
                "ingresos": sum(i["importe"] for i in data["ingresos"]),
                "gastos": sum(g["importe"] for g in data["gastos"])
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



# ==========================================
# PROYECTO INFORME DE GESTIÓN (SNAPSHOTS)
# ==========================================

class InformeAction(BaseModel):
    unidad_negocio: str
    periodo: str
    usuario: str

@app.get("/api/informes/estado")
def get_informe_estado(unidad_negocio: str, periodo: str):
    try:
        periodo = normalize_periodo(periodo)
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("""
            SELECT estado, usuario_apertura, fecha_apertura, usuario_cierre, fecha_cierre 
            FROM cert_informes_proyecto 
            WHERE unidad_negocio = %s AND periodo = %s
        """, (unidad_negocio, periodo))
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if row:
            return {
                "existe": True,
                "estado": row[0],
                "usuario_apertura": row[1],
                "fecha_apertura": row[2].isoformat() if row[2] else None,
                "usuario_cierre": row[3],
                "fecha_cierre": row[4].isoformat() if row[4] else None
            }
        else:
            return {
                "existe": False,
                "estado": "NO_INICIADO"
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/informes/iniciar")
def iniciar_informe(action: InformeAction):
    try:
        action.periodo = normalize_periodo(action.periodo)
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cert_informes_proyecto (unidad_negocio, periodo, estado, usuario_apertura)
            VALUES (%s, %s, 'ABIERTO', %s)
            ON CONFLICT (unidad_negocio, periodo) DO NOTHING
        """, (action.unidad_negocio, action.periodo, action.usuario))
        conn.commit()
        cur.close()
        conn.close()
        return {"message": "Informe iniciado correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/informes/cerrar")
def cerrar_informe(action: InformeAction):
    try:
        action.periodo = normalize_periodo(action.periodo)
        # Extraer snapshot de todos los módulos!
        # Usamos calculo vivo que ya incluye ingresos, gastos con RRHH, etc.
        vivo_data = get_informe_mensual_calculo_vivo(action.unidad_negocio, action.periodo)
        ingresos_data = vivo_data["ingresos"]
        gastos_data = vivo_data["gastos"]
        
        rrhh_data = get_rrhh(action.unidad_negocio, action.periodo)
        asientos_data = get_asientos(action.unidad_negocio, action.periodo)
        
        # Obtener datos de nuevos módulos para el snapshot
        try: consumos_data = get_consumos_inventarios_live(action.unidad_negocio, action.periodo)
        except: consumos_data = []
        
        try: equipos_data = get_equipos_live(action.unidad_negocio, action.periodo)
        except: equipos_data = []
        
        try: obras_data = get_certificados_obras_live(action.unidad_negocio, action.periodo)
        except: obras_data = []
        
        totales_data = {
            "ingresos": sum(i["importe"] for i in ingresos_data),
            "gastos": sum(g["importe"] for g in gastos_data),
            "neto": sum(i["importe"] for i in ingresos_data) - sum(g["importe"] for g in gastos_data)
        }

        snapshot = {
            "ingresos": ingresos_data,
            "rrhh": rrhh_data,
            "gastos": gastos_data,
            "asientos": asientos_data,
            "totales": totales_data,
            "consumos": consumos_data,
            "equipos": equipos_data,
            "obras": obras_data
        }

        import json
        snapshot_json = json.dumps(snapshot)

        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("""
            UPDATE cert_informes_proyecto
            SET estado = 'CERRADO',
                usuario_cierre = %s,
                fecha_cierre = CURRENT_TIMESTAMP,
                snapshot_data = %s
            WHERE unidad_negocio = %s AND periodo = %s
        """, (action.usuario, snapshot_json, action.unidad_negocio, action.periodo))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            raise Exception("No se encontró el informe abierto. Primero inícielo.")
            
        # También insertar en cert_cierres_mensuales y cert_cierres_detalle para retrocompatibilidad
        # Borrar previo si existiera
        cur.execute("SELECT id FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s", (action.unidad_negocio, action.periodo))
        old_cierre = cur.fetchone()
        if old_cierre:
            cur.execute("DELETE FROM cert_cierres_detalle WHERE cierre_id = %s", (old_cierre[0],))
            cur.execute("DELETE FROM cert_cierres_mensuales WHERE id = %s", (old_cierre[0],))
            
        cur.execute(
            "INSERT INTO cert_cierres_mensuales (unidad_negocio, periodo, usuario_cierre, fecha_cierre) VALUES (%s, %s, %s, NOW()) RETURNING id",
            (action.unidad_negocio, action.periodo, action.usuario)
        )
        cierre_id = cur.fetchone()[0]
        
        for item in ingresos_data + gastos_data:
            cur.execute(
                "INSERT INTO cert_cierres_detalle (cierre_id, origen, tipo_movimiento, categoria, fecha, concepto, comprobante, proveedor, importe) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (cierre_id, item["origen"], item["tipo_movimiento"], item["categoria"], item["fecha"] if item["fecha"] else None, item["concepto"], item["comprobante"], item.get("proveedor"), item["importe"])
            )
            
        conn.commit()
        cur.close()
        conn.close()
        return {"message": "Informe cerrado correctamente. Snapshot guardado."}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/informes/reabrir")
def reabrir_informe(action: InformeAction):
    try:
        action.periodo = normalize_periodo(action.periodo)
        # Aqui se podria verificar si el usuario es Admin.
        # Por ahora lo controlaremos en Frontend (solo Admin ve el boton).
        conn = get_supabase()
        cur = conn.cursor()
        cur.execute("""
            UPDATE cert_informes_proyecto
            SET estado = 'ABIERTO',
                usuario_cierre = NULL,
                fecha_cierre = NULL,
                snapshot_data = NULL
            WHERE unidad_negocio = %s AND periodo = %s
        """, (action.unidad_negocio, action.periodo))
        
        # También borrar de cert_cierres_mensuales y cert_cierres_detalle
        cur.execute("SELECT id FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s", (action.unidad_negocio, action.periodo))
        row = cur.fetchone()
        if row:
            cierre_id = row[0]
            cur.execute("DELETE FROM cert_cierres_detalle WHERE cierre_id = %s", (cierre_id,))
            cur.execute("DELETE FROM cert_cierres_mensuales WHERE id = %s", (cierre_id,))

        conn.commit()
        cur.close()
        conn.close()
        return {"message": "Informe reabierto correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/informes/eliminar")
def eliminar_informe(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    try:
        periodo = normalize_periodo(periodo)
        conn = get_supabase()
        cur = conn.cursor()
        
        # Verificar que el informe exista y esté ABIERTO
        cur.execute("""
            SELECT estado FROM cert_informes_proyecto 
            WHERE unidad_negocio = %s AND periodo = %s
        """, (unidad_negocio, periodo))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            raise HTTPException(status_code=404, detail="El informe no existe.")
            
        estado = row[0]
        if estado != 'ABIERTO':
            cur.close()
            conn.close()
            raise HTTPException(status_code=400, detail="Solo se pueden eliminar informes en estado ABIERTO.")
            
        # Eliminar el informe
        cur.execute("""
            DELETE FROM cert_informes_proyecto 
            WHERE unidad_negocio = %s AND periodo = %s
        """, (unidad_negocio, periodo))
        
        # Eliminar cierres del mismo periodo y unidad por precaución
        cur.execute("SELECT id FROM cert_cierres_mensuales WHERE unidad_negocio = %s AND periodo = %s", (unidad_negocio, periodo))
        row_c = cur.fetchone()
        if row_c:
            cierre_id = row_c[0]
            cur.execute("DELETE FROM cert_cierres_detalle WHERE cierre_id = %s", (cierre_id,))
            cur.execute("DELETE FROM cert_cierres_mensuales WHERE id = %s", (cierre_id,))

        conn.commit()
        cur.close()
        conn.close()
        return {"message": "Informe eliminado correctamente"}
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/informes/lista")
def get_informes_lista(user=Depends(get_current_user)):
    try:
        conn = get_supabase()
        cur = conn.cursor()
        
        # Filtramos por las unidades a las que tiene acceso el usuario
        unidades_permitidas = set()
        if user.get("rol") != "admin":
            cur.execute("SELECT unidad_negocio FROM cert_usuarios_unidades WHERE usuario_id = %s", (user.get('id') or user.get('sub'),))
            unidades_permitidas = {r[0].strip() for r in cur.fetchall()}
            if not unidades_permitidas:
                cur.close()
                conn.close()
                return []
                
        if user.get("rol") == "admin":
            cur.execute("SELECT id, unidad_negocio, periodo, estado, usuario_apertura FROM cert_informes_proyecto ORDER BY periodo DESC")
        else:
            cur.execute("SELECT id, unidad_negocio, periodo, estado, usuario_apertura FROM cert_informes_proyecto WHERE unidad_negocio IN %s ORDER BY periodo DESC", (tuple(unidades_permitidas),))
            
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        result = []
        for r in rows:
            result.append({
                "id": r[0],
                "unidad_negocio": r[1],
                "periodo": r[2],
                "estado": r[3],
                "usuario_apertura": r[4]
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/respaldos")
def listar_respaldos(tipo_documento: str, unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    periodo = normalize_periodo(periodo)
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, nombre_archivo, tipo_mime, usuario_carga, fecha_carga
            FROM cert_respaldos
            WHERE tipo_documento = %s AND unidad_negocio = %s AND periodo = %s
            ORDER BY fecha_carga DESC
        """, (tipo_documento, unidad_negocio, periodo))
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                "id": r[0],
                "nombre_archivo": r[1],
                "tipo_mime": r[2],
                "usuario_carga": r[3],
                "fecha_carga": str(r[4])
            })
        return result
    finally:
        cur.close()
        conn.close()

@app.post("/api/respaldos/upload")
async def upload_respaldo(
    tipo_documento: str = Form(...),
    unidad_negocio: str = Form(...),
    periodo: str = Form(...),
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    periodo = normalize_periodo(periodo)
    import base64
    contents = await file.read()
    encoded = base64.b64encode(contents).decode("utf-8")
    
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO cert_respaldos (tipo_documento, unidad_negocio, periodo, nombre_archivo, tipo_mime, contenido, usuario_carga)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (tipo_documento, unidad_negocio, periodo, file.filename, file.content_type, encoded, current_user.get("email", "unknown")))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return {"status": "ok", "message": "Archivo adjunto guardado correctamente"}

@app.get("/api/respaldos/descargar/{id}")
def descargar_respaldo(id: int, current_user = Depends(get_current_user)):
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("SELECT nombre_archivo, tipo_mime, contenido FROM cert_respaldos WHERE id = %s", (id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        nombre, mime, encoded = row
        import base64
        from fastapi import Response
        decoded = base64.b64decode(encoded)
        headers = {
            "Content-Disposition": f"attachment; filename={nombre}"
        }
        return Response(content=decoded, media_type=mime, headers=headers)
    finally:
        cur.close()
        conn.close()

@app.delete("/api/respaldos/{id}")
def eliminar_respaldo(id: int, current_user = Depends(get_current_user)):
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM cert_respaldos WHERE id = %s", (id,))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return {"status": "ok", "message": "Archivo adjunto eliminado"}

# ═══════════════════════════════════════════════════════
# NUEVOS MÓDULOS: CONSUMOS, EQUIPOS Y OBRAS
# ═══════════════════════════════════════════════════════

def get_consumos_inventarios_live(unidad_negocio: str, periodo: str):
    periodo = normalize_periodo(periodo)
    conn = get_aurora()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT 
                fecha, 
                numerocomprobante, 
                productoconsumoprod, 
                cantidadconsumoprod, 
                unidadconsumoprod, 
                preciounitvalorizadoconsumoprod, 
                importevalorizadoconsumoprod,
                ordendeproduccion,
                depositoorigenconsumoprod,
                empresa
            FROM analisis_de_consumos_de_produccion
            WHERE (empresa = %s OR depositoorigenconsumoprod ILIKE %s OR depositoorigenconsumoprod = %s)
              AND SUBSTRING(fecha, 1, 7) = %s
            ORDER BY fecha DESC
        """, (unidad_negocio, f"%{unidad_negocio}%", unidad_negocio, periodo))
        rows = cur.fetchall()
        
        result = []
        for r in rows:
            raw_precio = r[5]
            raw_importe = r[6]
            
            precio = 0.0
            if raw_precio and raw_precio != 'NULL' and raw_precio != '':
                try: precio = float(raw_precio)
                except: pass
                
            importe = 0.0
            if raw_importe and raw_importe != 'NULL' and raw_importe != '':
                try: importe = float(raw_importe)
                except: pass
                
            result.append({
                "fecha": str(r[0])[:10] if r[0] else "",
                "comprobante": r[1] or "",
                "insumo": r[2] or "",
                "cantidad": float(r[3] or 0.0),
                "unidad": r[4] or "",
                "precio_unitario": precio,
                "total": importe,
                "orden_produccion": r[7] or "",
                "deposito": r[8] or "",
                "sucursal": r[9] or ""
            })
        return result
    finally:
        cur.close()
        conn.close()

@app.get("/api/consumos-inventarios")
def get_consumos_inventarios(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    periodo = normalize_periodo(periodo)
    cerrado = check_informe_cerrado(unidad_negocio, periodo, 'consumos')
    if cerrado is not None:
        return cerrado
    try:
        return get_consumos_inventarios_live(unidad_negocio, periodo)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def get_certificaciones_maquinas_new(unidad_negocio: str, periodo: str):
    import urllib.request
    import json
    import calendar
    import datetime
    
    periodo = normalize_periodo(periodo)
    y, m = periodo.split('-')
    y_int, m_int = int(y), int(m)
    
    # 1. Fetch configurations from local database overlapping with this period
    # Period start and end dates
    start_date = datetime.date(y_int, m_int, 1)
    _, last_day = calendar.monthrange(y_int, m_int)
    end_date = datetime.date(y_int, m_int, last_day)
    
    assigned_configs = []
    try:
        conn_local = get_supabase()
        cur_local = conn_local.cursor()
        cur_local.execute("""
            SELECT id, sucursal, equipo_codigo, equipo_nombre, fecha_desde, fecha_hasta
            FROM cert_config_equipos_asignados
            WHERE fecha_desde <= %s AND (fecha_hasta IS NULL OR fecha_hasta >= %s)
        """, (end_date, start_date))
        rows = cur_local.fetchall()
        cur_local.close()
        conn_local.close()
        for r in rows:
            assigned_configs.append({
                "id": r[0],
                "sucursal": r[1],
                "equipo_codigo": r[2],
                "equipo_nombre": r[3],
                "fecha_desde": r[4],
                "fecha_hasta": r[5]
            })
    except Exception as db_err:
        print("Error fetching cert_config_equipos_asignados:", db_err)
        
    # Maps equipo_nombre.lower().strip() -> configured sucursal
    # (or mapping using code, we can check both name and code)
    name_to_config = {}
    code_to_config = {}
    
    for cfg in assigned_configs:
        eq_name = str(cfg["equipo_nombre"]).strip().lower()
        eq_code = str(cfg["equipo_codigo"]).strip().lower()
        name_to_config[eq_name] = cfg
        code_to_config[eq_code] = cfg

    # 2. Fetch original certifications from Supabase
    url = f"https://naxjzquhdzyoxtjataaw.supabase.co/rest/v1/certificaciones_maquinas?anio=eq.{y_int}&mes=eq.{m_int}"
    apikey = "sb_publishable_K3r5ogAi25D2ffI6-9cWLg_VJddbdxF"
    
    req = urllib.request.Request(
        url,
        headers={
            "apikey": apikey,
            "Authorization": f"Bearer {apikey}"
        }
    )
    
    supabase_data = []
    try:
        with urllib.request.urlopen(req, timeout=5) as response:
            supabase_data = json.loads(response.read().decode('utf-8'))
    except Exception as e:
        print("Error fetching certificaciones_maquinas from Supabase:", e)

    # Process and override original records
    final_certs = []
    processed_configs = set()
    
    un_req = unidad_negocio.strip().lower().replace("cee", "").replace("enriquez", "").strip()
    
    for r in supabase_data:
        maquina = str(r.get('maquina') or '').strip()
        maquina_clean = maquina.lower()
        maquina_code_clean = maquina_clean.split('/')[0].strip()
        
        cfg = name_to_config.get(maquina_clean) or code_to_config.get(maquina_code_clean)
        
        if cfg:
            r['unidad_de_negocio'] = cfg['sucursal']
            r['observaciones'] = (r.get('observaciones') or '') + " [Asignación configurada]"
            processed_configs.add(cfg['id'])
            
        final_certs.append(r)
        
    # Create virtual records for any configured assignments that were NOT in the Supabase response
    for cfg in assigned_configs:
        if cfg['id'] not in processed_configs:
            virtual_row = {
                "id": -200 - cfg['id'],
                "maquina": cfg['equipo_nombre'],
                "unidad_de_negocio": cfg['sucursal'],
                "mes": m_int,
                "anio": y_int,
                "observaciones": "Asignación configurada (Sin certificación en Supabase)",
                "created_at": datetime.datetime.now().isoformat()
            }
            final_certs.append(virtual_row)

    # Now filter by requested unidad_negocio
    filtered = []
    for r in final_certs:
        un_db = str(r.get('unidad_de_negocio') or '').strip().lower().replace("cee", "").replace("enriquez", "").strip()
        if un_req == un_db or un_req in un_db or un_db in un_req:
            filtered.append(r)
            
    if not filtered and not assigned_configs:
        mock_1 = {
            "id": -101,
            "maquina": "AE0001/Autoelevador",
            "mes": m_int,
            "anio": y_int,
            "observaciones": "Certificación de prueba (Autogenerada porque la tabla está vacía)",
            "unidad_de_negocio": unidad_negocio,
            "created_at": "2026-06-21T18:00:00Z"
        }
        mock_2 = {
            "id": -102,
            "maquina": "M0169/Camioneta",
            "mes": m_int,
            "anio": y_int,
            "observaciones": "Certificación mensual de camioneta taller",
            "unidad_de_negocio": unidad_negocio,
            "created_at": "2026-06-21T18:00:00Z"
        }
        filtered = [mock_1, mock_2]
        
    return filtered

def get_equipos_live(unidad_negocio: str, periodo: str):
    periodo = normalize_periodo(periodo)
    # 1. Planilla
    conn_supa = get_supabase()
    cur_supa = conn_supa.cursor()
    cur_supa.execute("""
        SELECT id, equipo, concepto, horas_kilometros, precio_unitario, total, usuario_carga, fecha_carga
        FROM cert_equipos_planilla
        WHERE unidad_negocio = %s AND periodo = %s
    """, (unidad_negocio, periodo))
    rows_supa = cur_supa.fetchall()
    cur_supa.close()
    conn_supa.close()
    
    result = []
    for r in rows_supa:
        result.append({
            "id": r[0],
            "origen": "PLANILLA",
            "equipo": r[1],
            "concepto": r[2],
            "horas_kilometros": float(r[3] or 0),
            "precio_unitario": float(r[4] or 0),
            "total": float(r[5] or 0),
            "usuario_carga": r[6],
            "fecha_carga": str(r[7])
        })
        
    # 2. Finnegans
    y, m = periodo.split('-')
    conn_a = get_aurora()
    cur_a = conn_a.cursor()
    try:
        cur_a.execute("""
            SELECT fecha, documento, comprobante, equiposolicitantenombre, productonombre, itemcantidad, itemprecio, itemimporte, conceptonombre, descripcion
            FROM ceesa_cee_certificados_ventas_internas
            WHERE empresa = 'Taller Central CEE Enriquez'
              AND EXTRACT(YEAR FROM CAST(fecha AS TIMESTAMP)) = %s
              AND EXTRACT(MONTH FROM CAST(fecha AS TIMESTAMP)) = %s
        """, (int(y), int(m)))
        rows_a = cur_a.fetchall()
        
        for r in rows_a:
            eq_solicitante = r[3] or ""
            eq_clean = eq_solicitante.strip().lower()
            unit_clean = unidad_negocio.strip().lower()
            
            matches = (eq_clean == unit_clean or eq_clean in unit_clean or unit_clean in eq_clean)
            if matches:
                qty = 0.0
                try: qty = float(r[5] or 0)
                except: pass
                
                price = 0.0
                try: price = float(r[6] or 0)
                except: pass
                
                total = 0.0
                try: total = float(r[7] or 0)
                except: pass
                
                result.append({
                    "id": None,
                    "origen": "FINNEGANS",
                    "equipo": r[4] or "Equipo sin nombre",
                    "concepto": r[8] or r[9] or "Alquiler",
                    "horas_kilometros": qty,
                    "precio_unitario": price,
                    "total": total,
                    "documento": r[1] or "",
                    "comprobante": r[2] or "",
                    "fecha": str(r[0])[:10] if r[0] else ""
                })
    finally:
        cur_a.close()
        conn_a.close()
        
    # 3. Supabase Nuevo (certificaciones_maquinas)
    un_req = unidad_negocio.strip().lower().replace("cee", "").replace("enriquez", "").strip()
    supabase_certs = get_certificaciones_maquinas_new(unidad_negocio, periodo)
    if supabase_certs:
        conn_a = get_aurora()
        cur_a = conn_a.cursor()
        try:
            # Query work logs in Finnegans for this period
            cur_a.execute("""
                SELECT maquina, codigomaquina, fecha, horastrabajadas, descripciontrabajo, centrodecosto, maquinista
                FROM ceesa_cee_equipos_trabajos_realizados
                WHERE EXTRACT(YEAR FROM CAST(fecha AS TIMESTAMP)) = %s
                  AND EXTRACT(MONTH FROM CAST(fecha AS TIMESTAMP)) = %s
            """, (int(y), int(m)))
            all_work_logs = cur_a.fetchall()
            
            # Query billing prices for this period
            cur_a.execute("""
                SELECT productonombre, itemprecio
                FROM ceesa_cee_certificados_ventas_internas
                WHERE empresa = 'Taller Central CEE Enriquez'
                  AND EXTRACT(YEAR FROM CAST(fecha AS TIMESTAMP)) = %s
                  AND EXTRACT(MONTH FROM CAST(fecha AS TIMESTAMP)) = %s
            """, (int(y), int(m)))
            all_prices = cur_a.fetchall()
            
            price_map = {}
            for row in all_prices:
                p_name = str(row[0] or "").strip().lower()
                try:
                    price_map[p_name] = float(row[1] or 0)
                except:
                    pass
                    
            for cert in supabase_certs:
                maquina_cert = str(cert.get('maquina') or '').strip()
                maquina_cert_clean = maquina_cert.split('/')[0].strip().lower()
                
                # Match work logs
                total_hours = 0.0
                trabajos_set = set()
                maquinistas_set = set()
                
                for wl in all_work_logs:
                    wl_maquina = str(wl[0] or '').strip().lower()
                    wl_codigo = str(wl[1] or '').strip().lower()
                    wl_cc = str(wl[5] or '').strip().lower()
                    
                    m_match = (maquina_cert_clean == wl_codigo or 
                               maquina_cert_clean in wl_maquina or 
                               wl_codigo in maquina_cert_clean)
                    
                    cc_clean = wl_cc.replace("cee", "").replace("enriquez", "").strip()
                    cc_match = (un_req == cc_clean or un_req in cc_clean or cc_clean in un_req)
                    
                    if m_match and cc_match:
                        try:
                            total_hours += float(wl[3] or 0)
                        except:
                            pass
                        if wl[4] and str(wl[4]).strip() and str(wl[4]).strip().lower() != 'null':
                            trabajos_set.add(str(wl[4]).strip())
                        if wl[6] and str(wl[6]).strip() and str(wl[6]).strip().lower() != 'null':
                            maquinistas_set.add(str(wl[6]).strip())
                            
                # Determine unit price (default to a historical price or 0 if not found)
                price_unit = 0.0
                for p_name, val in price_map.items():
                    if maquina_cert_clean in p_name or p_name in maquina_cert.lower():
                        price_unit = val
                        break
                        
                # Fallback prices for testing (e.g. AE0001 is 25000, M0169 is 12000)
                if price_unit == 0.0:
                    if 'ae0001' in maquina_cert_clean:
                        price_unit = 25000.0
                    elif 'm0169' in maquina_cert_clean:
                        price_unit = 15000.0
                    else:
                        price_unit = 18000.0
                        
                result.append({
                    "id": cert.get('id'),
                    "origen": "SUPABASE_CERT",
                    "equipo": maquina_cert,
                    "concepto": cert.get('observaciones') or "Certificación Imputada",
                    "horas_kilometros": total_hours if total_hours > 0 else 8.0,
                    "precio_unitario": price_unit,
                    "total": (total_hours if total_hours > 0 else 8.0) * price_unit,
                    "documento": "Supabase Cert",
                    "comprobante": "",
                    "fecha": f"{y}-{m}-01",
                    "usuario_carga": "Supabase",
                    "fecha_carga": cert.get('created_at') or "",
                    "detalles_trabajos": list(trabajos_set),
                    "operarios": list(maquinistas_set),
                    "unidad_de_negocio": cert.get('unidad_de_negocio'),
                    "mes": cert.get('mes'),
                    "anio": cert.get('anio')
                })
        except Exception as e:
            print("Error joining certificaciones_maquinas with Finnegans:", e)
        finally:
            cur_a.close()
            conn_a.close()
            
    return result

@app.get("/api/equipos")
def get_equipos(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    periodo = normalize_periodo(periodo)
    cerrado = check_informe_cerrado(unidad_negocio, periodo, 'equipos')
    if cerrado is not None:
        return cerrado
    try:
        return get_equipos_live(unidad_negocio, periodo)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/equipos/upload")
async def upload_equipos(
    file: UploadFile = File(...),
    unidad_negocio: str = Form(...),
    periodo: str = Form(...),
    current_user = Depends(get_current_user)
):
    periodo = normalize_periodo(periodo)
    import io
    import openpyxl
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    header_row_idx = None
    headers = []
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row: continue
        row_str = [str(c or "").strip().lower() for c in row]
        if any("equipo" in s for s in row_str) and any("concepto" in s for s in row_str):
            header_row_idx = row_idx
            headers = row_str
            break
            
    if header_row_idx is None:
        headers = [str(c or "").strip().lower() for c in ws[1]]
        header_row_idx = 1
        
    idx_equipo = -1
    idx_concepto = -1
    idx_horas = -1
    idx_precio = -1
    idx_total = -1
    
    for i, h in enumerate(headers):
        if "equipo" in h: idx_equipo = i
        elif "concepto" in h: idx_concepto = i
        elif "hora" in h or "kilom" in h or "km" in h: idx_horas = i
        elif "precio" in h or "unit" in h: idx_precio = i
        elif "total" in h or "importe" in h: idx_total = i
        
    if idx_equipo == -1 or idx_concepto == -1:
        raise HTTPException(status_code=400, detail="El archivo no tiene las columnas requeridas (Equipo, Concepto).")
        
    conn = get_supabase()
    cur = conn.cursor()
    user_email = current_user.get("email", "unknown")
    
    inserted_count = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row or not any(row): continue
        
        equipo_val = str(row[idx_equipo] or "").strip()
        concepto_val = str(row[idx_concepto] or "").strip()
        
        if not equipo_val: continue
        
        horas_val = 0.0
        if idx_horas != -1 and row[idx_horas] is not None:
            try: horas_val = float(row[idx_horas])
            except: pass
            
        precio_val = 0.0
        if idx_precio != -1 and row[idx_precio] is not None:
            try: precio_val = float(row[idx_precio])
            except: pass
            
        total_val = horas_val * precio_val
        if idx_total != -1 and row[idx_total] is not None:
            try: total_val = float(row[idx_total])
            except: pass
            
        try:
            cur.execute("""
                INSERT INTO cert_equipos_planilla (unidad_negocio, periodo, equipo, concepto, horas_kilometros, precio_unitario, total, usuario_carga)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (unidad_negocio, periodo, equipo_val, concepto_val, horas_val, precio_val, total_val, user_email))
            inserted_count += 1
        except Exception as e:
            errors.append(f"Fila {row_idx}: {str(e)}")
            
    conn.commit()
    cur.close()
    conn.close()
    
    return {"status": "ok", "inserted": inserted_count, "errors": errors}

@app.delete("/api/equipos/bulk")
def delete_equipos_bulk(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    periodo = normalize_periodo(periodo)
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM cert_equipos_planilla WHERE unidad_negocio = %s AND periodo = %s", (unidad_negocio, periodo))
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/api/equipos/{id}")
def delete_equipo_item(id: int, current_user = Depends(get_current_user)):
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM cert_equipos_planilla WHERE id = %s", (id,))
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


def get_certificados_obras_live(unidad_negocio: str, periodo: str):
    periodo = normalize_periodo(periodo)
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, numero_interno, comitente, contratista, obra, fecha_certificado, estado, usuario_carga, fecha_carga
            FROM cert_obras_maestro
            WHERE unidad_negocio = %s AND periodo = %s
            ORDER BY numero_interno ASC
        """, (unidad_negocio, periodo))
        maestros = cur.fetchall()
        
        sheets = []
        for m in maestros:
            maestro_id = m[0]
            cur.execute("""
                SELECT id, item, descripcion, unidad_medida, cantidad_aprobada, precio_unitario, presente_certificado, anterior_certificado, total_certificado, faltante_certificar, parcial_presente, parcial_anterior, parcial_total, monto_aprobado, avance_usd
                FROM cert_obras_detalles
                WHERE maestro_id = %s
                ORDER BY id ASC
            """, (maestro_id,))
            details = cur.fetchall()
            
            items_list = []
            for d in details:
                items_list.append({
                    "id": d[0],
                    "item": d[1] or "",
                    "descripcion": d[2] or "",
                    "unidad_medida": d[3] or "",
                    "cantidad_aprobada": float(d[4] or 0),
                    "precio_unitario": float(d[5] or 0),
                    "presente_certificado": float(d[6] or 0),
                    "anterior_certificado": float(d[7] or 0),
                    "total_certificado": float(d[8] or 0),
                    "faltante_certificar": float(d[9] or 0),
                    "parcial_presente": float(d[10] or 0),
                    "parcial_anterior": float(d[11] or 0),
                    "parcial_total": float(d[12] or 0),
                    "monto_aprobado": float(d[13] or 0),
                    "avance_usd": float(d[14] or 0)
                })
                
            sheets.append({
                "id": maestro_id,
                "numero_interno": m[1],
                "comitente": m[2] or "",
                "contratista": m[3] or "",
                "obra": m[4] or "",
                "fecha_certificado": str(m[5]) if m[5] else "",
                "estado": m[6],
                "usuario_carga": m[7],
                "fecha_carga": str(m[8]),
                "items": items_list
            })
        return sheets
    finally:
        cur.close()
        conn.close()

@app.get("/api/certificados-obras")
def get_certificados_obras(unidad_negocio: str, periodo: str, current_user = Depends(get_current_user)):
    periodo = normalize_periodo(periodo)
    cerrado = check_informe_cerrado(unidad_negocio, periodo, 'obras')
    if cerrado is not None:
        return cerrado
    try:
        return get_certificados_obras_live(unidad_negocio, periodo)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/certificados-obras/upload")
async def upload_certificado_obra(
    file: UploadFile = File(...),
    unidad_negocio: str = Form(...),
    periodo: str = Form(...),
    current_user = Depends(get_current_user)
):
    periodo = normalize_periodo(periodo)
    import io
    import openpyxl
    from datetime import datetime, date
    
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    comitente = ""
    contratista = ""
    obra = ""
    fecha_cert = None
    
    for row in ws.iter_rows(max_row=15, values_only=True):
        if not row: continue
        for col_idx, cell in enumerate(row):
            cell_str = str(cell or "").strip().lower()
            if "comitente" in cell_str and col_idx + 1 < len(row):
                comitente = str(row[col_idx + 1] or "").strip()
            elif "contratista" in cell_str and col_idx + 1 < len(row):
                contratista = str(row[col_idx + 1] or "").strip()
            elif "obra" in cell_str and col_idx + 1 < len(row):
                if "total" not in cell_str:
                    obra = str(row[col_idx + 1] or "").strip()
            elif "fecha" in cell_str and col_idx + 1 < len(row):
                val = row[col_idx + 1]
                if isinstance(val, datetime):
                    fecha_cert = val.date()
                elif isinstance(val, date):
                    fecha_cert = val
                elif val:
                    try:
                        fecha_cert = datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
                    except:
                        try:
                            fecha_cert = datetime.strptime(str(val).strip(), "%d/%m/%Y").date()
                        except:
                            pass

    header_row_idx = None
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row: continue
        row_str = [str(c or "").strip().lower() for c in row]
        if any("item" in s for s in row_str) and any("descrip" in s for s in row_str) and any("unidad" in s for s in row_str):
            header_row_idx = row_idx
            headers = row_str
            break
            
    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="No se encontró la fila de cabecera con las columnas (Item, Descripción, Unidad de Medida).")
        
    idx_item = -1
    idx_desc = -1
    idx_um = -1
    idx_cant_aprob = -1
    idx_precio = -1
    idx_pres_cert = -1
    idx_ant_cert = -1
    idx_tot_cert = -1
    idx_faltante = -1
    idx_parc_pres = -1
    idx_parc_ant = -1
    idx_parc_tot = -1
    idx_monto_aprob = -1
    idx_avance = -1
    
    for i, h in enumerate(headers):
        if "item" in h: idx_item = i
        elif "descrip" in h or "tarea" in h or "trabajo" in h: idx_desc = i
        elif "unidad" in h or "u.m" in h: idx_um = i
        elif "aprobada" in h or "cantidad aprob" in h: idx_cant_aprob = i
        elif "precio" in h or "unitario" in h: idx_precio = i
        elif "presente" in h and "certificado" in h and "parcial" not in h: idx_pres_cert = i
        elif "anterior" in h and "certificado" in h and "parcial" not in h: idx_ant_cert = i
        elif "total" in h and "certificado" in h and "parcial" not in h: idx_tot_cert = i
        elif "faltante" in h: idx_faltante = i
        elif "parcial presente" in h or ("parcial" in h and "presente" in h): idx_parc_pres = i
        elif "parcial anterior" in h or ("parcial" in h and "anterior" in h): idx_parc_ant = i
        elif "parcial total" in h or ("parcial" in h and "total" in h): idx_parc_tot = i
        elif "monto aprobado" in h or "aprobado" in h: idx_monto_aprob = i
        elif "avance" in h: idx_avance = i
        
    if idx_item == -1 or idx_desc == -1:
         raise HTTPException(status_code=400, detail="Las columnas básicas 'Item' o 'Descripción' no fueron detectadas.")
         
    conn = get_supabase()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT COALESCE(MAX(numero_interno), 0) + 1 
        FROM cert_obras_maestro 
        WHERE unidad_negocio = %s AND periodo = %s
    """, (unidad_negocio, periodo))
    next_num = cur.fetchone()[0]
    
    user_email = current_user.get("email", "unknown")
    
    cur.execute("""
        INSERT INTO cert_obras_maestro (unidad_negocio, periodo, numero_interno, comitente, contratista, obra, fecha_certificado, estado, usuario_carga)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'BORRADOR', %s) RETURNING id
    """, (unidad_negocio, periodo, next_num, comitente, contratista, obra, fecha_cert, user_email))
    maestro_id = cur.fetchone()[0]
    
    def get_cell_val(row, idx):
        if idx != -1 and idx < len(row) and row[idx] is not None:
            val = row[idx]
            if str(val).strip().upper() == "NULL" or str(val).strip() == "":
                return 0.0
            try: return float(val)
            except: return 0.0
        return 0.0
        
    inserted_items = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not row or not any(row): continue
        item_val = str(row[idx_item] or "").strip()
        desc_val = str(row[idx_desc] or "").strip()
        
        if not item_val and not desc_val: continue
        
        um_val = str(row[idx_um] or "").strip() if idx_um != -1 else ""
        
        cant_aprob = get_cell_val(row, idx_cant_aprob)
        precio_unit = get_cell_val(row, idx_precio)
        pres_cert = get_cell_val(row, idx_pres_cert)
        ant_cert = get_cell_val(row, idx_ant_cert)
        
        tot_cert = get_cell_val(row, idx_tot_cert) if idx_tot_cert != -1 else 0.0
        if idx_tot_cert == -1 or tot_cert == 0.0:
            tot_cert = pres_cert + ant_cert
            
        faltante = get_cell_val(row, idx_faltante) if idx_faltante != -1 else 0.0
        if idx_faltante == -1 or faltante == 0.0:
            faltante = cant_aprob - tot_cert
            
        parc_pres = get_cell_val(row, idx_parc_pres) if idx_parc_pres != -1 else 0.0
        if idx_parc_pres == -1 or parc_pres == 0.0:
            parc_pres = pres_cert * precio_unit
            
        parc_ant = get_cell_val(row, idx_parc_ant) if idx_parc_ant != -1 else 0.0
        if idx_parc_ant == -1 or parc_ant == 0.0:
            parc_ant = ant_cert * precio_unit
            
        parc_tot = get_cell_val(row, idx_parc_tot) if idx_parc_tot != -1 else 0.0
        if idx_parc_tot == -1 or parc_tot == 0.0:
            parc_tot = tot_cert * precio_unit
            
        monto_aprob = get_cell_val(row, idx_monto_aprob) if idx_monto_aprob != -1 else 0.0
        if idx_monto_aprob == -1 or monto_aprob == 0.0:
            monto_aprob = cant_aprob * precio_unit
            
        avance = get_cell_val(row, idx_avance) if idx_avance != -1 else 0.0
        if idx_avance == -1 or avance == 0.0:
            avance = (tot_cert / cant_aprob * 100.0) if cant_aprob > 0.0 else 0.0
        
        cur.execute("""
            INSERT INTO cert_obras_detalles (
                maestro_id, item, descripcion, unidad_medida, cantidad_aprobada, precio_unitario, 
                presente_certificado, anterior_certificado, total_certificado, faltante_certificar, 
                parcial_presente, parcial_anterior, parcial_total, monto_aprobado, avance_usd
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            maestro_id, item_val, desc_val, um_val, cant_aprob, precio_unit,
            pres_cert, ant_cert, tot_cert, faltante,
            parc_pres, parc_ant, parc_tot, monto_aprob, avance
        ))
        inserted_items += 1
        
    conn.commit()
    cur.close()
    conn.close()
    
    return {
        "status": "ok",
        "maestro_id": maestro_id,
        "numero_interno": next_num,
        "inserted_items": inserted_items
    }

class ConfirmCertificadoObraReq(BaseModel):
    comitente: str
    contratista: str
    obra: str
    fecha_certificado: Optional[str] = None
    items: list

@app.put("/api/certificados-obras/{maestro_id}/confirm")
def confirm_certificado_obra(maestro_id: int, req: ConfirmCertificadoObraReq, current_user = Depends(get_current_user)):
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE cert_obras_maestro 
            SET comitente = %s, contratista = %s, obra = %s, fecha_certificado = %s, estado = 'CONFIRMADO'
            WHERE id = %s
        """, (req.comitente, req.contratista, req.obra, req.fecha_certificado if req.fecha_certificado else None, maestro_id))
        
        cur.execute("DELETE FROM cert_obras_detalles WHERE maestro_id = %s", (maestro_id,))
        
        for item in req.items:
            cur.execute("""
                INSERT INTO cert_obras_detalles (
                    maestro_id, item, descripcion, unidad_medida, cantidad_aprobada, precio_unitario, 
                    presente_certificado, anterior_certificado, total_certificado, faltante_certificar, 
                    parcial_presente, parcial_anterior, parcial_total, monto_aprobado, avance_usd
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                maestro_id, item.get("item"), item.get("descripcion"), item.get("unidad_medida"),
                item.get("cantidad_aprobada", 0), item.get("precio_unitario", 0),
                item.get("presente_certificado", 0), item.get("anterior_certificado", 0),
                item.get("total_certificado", 0), item.get("faltante_certificar", 0),
                item.get("parcial_presente", 0), item.get("parcial_anterior", 0),
                item.get("parcial_total", 0), item.get("monto_aprobado", 0),
                item.get("avance_usd", 0)
            ))
            
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.delete("/api/certificados-obras/{maestro_id}")
def delete_certificado_obra(maestro_id: int, current_user = Depends(get_current_user)):
    conn = get_supabase()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM cert_obras_maestro WHERE id = %s", (maestro_id,))
        conn.commit()
        return {"status": "ok"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()
